"""
Northern Solar - DWG & Excel Extractor Engine
=============================================
Enjin automasi untuk:
1. Mengekstrak maklumat daripada Lukisan Teknikal DWG PDF (Title block, PV Layout, String config, Inverter, Bateri).
2. Mengekstrak maklumat projek daripada fail Excel (Template.xlsx / PAGE 1 - PAGE 7).
3. Mengisi nilai ke dalam sel-sel KUNING Template.xlsx.
4. Menjana semula Laporan Simulasi PVsyst PDF yang bersih (tanpa highlight kuning)
   mengikut template Battery atau No-Battery secara automatik.
"""

import os
import io
import re
import glob
import math
import openpyxl
from openpyxl.styles import PatternFill
import pymupdf

# ==============================================================================
# 1. KONFIGURASI RUJUKAN NEGERI & CUACA
# ==============================================================================

STATE_WEATHER_RATES = [
    ("PERLIS", "Perlis", 3.5),
    ("KEDAH", "Kedah", 3.5),
    ("PULAU PINANG", "Pulau Pinang", 3.8),
    ("PENANG", "Pulau Pinang", 3.8),
    ("PERAK", "Perak", 3.8),
    ("SELANGOR", "Selangor", 3.9),
    ("KUALA LUMPUR", "Kuala Lumpur", 3.9),
    ("KL", "Kuala Lumpur", 3.9),
    ("PUTRAJAYA", "Putrajaya", 3.9),
    ("NEGERI SEMBILAN", "Negeri Sembilan", 4.0),
    ("MELAKA", "Melaka", 4.0),
    ("MALACCA", "Melaka", 4.0),
    ("JOHOR", "Johor", 4.0),
    ("PAHANG", "Pahang", 4.2),
    ("TERENGGANU", "Terengganu", 4.3),
    ("KELANTAN", "Kelantan", 4.3),
    ("SABAH", "Sabah", 4.5),
    ("SARAWAK", "Sarawak", 4.5),
]

def get_weather_assumption_for_state(state_name):
    if not state_name:
        return 3.8
    s_up = state_name.upper()
    for st_key, _, rate in STATE_WEATHER_RATES:
        if st_key in s_up:
            return rate
    return 3.8

def get_year_input_for_state(state_name, seed_str=""):
    state_up = (state_name or "").upper()
    if any(s in state_up for s in ["PENANG", "PINANG", "KEDAH", "PERLIS", "PERAK", "UTARA"]):
        min_val, max_val = 1710, 1795
    elif any(s in state_up for s in ["JOHOR", "MELAKA", "SEMBILAN", "SELATAN"]):
        min_val, max_val = 1610, 1690
    else:
        min_val, max_val = 1660, 1740

    if seed_str:
        val_hash = sum(ord(c) * 3 for c in seed_str)
        diff = max_val - min_val
        val = min_val + (val_hash % (diff + 1))
        return round(val / 5) * 5
    return 1780 if any(s in state_up for s in ["PENANG", "PINANG", "KEDAH", "PERLIS", "PERAK"]) else 1680

def get_annual_load_for_model(inverter_model, seed_str="", system_size_kwp=0.0):
    inv_up = (inverter_model or "").upper()
    is_hit = "HIT" in inv_up or "THREE" in inv_up or system_size_kwp >= 6.0
    if is_hit:
        min_val, max_val = 8800, 13500
    else:
        min_val, max_val = 5500, 7800

    if seed_str:
        val_hash = sum(ord(c) * 7 for c in seed_str)
        diff = max_val - min_val
        val = min_val + (val_hash % (diff + 1))
        return round(val / 50) * 50
    return 9600 if is_hit else 6500

# ==============================================================================
# 2. ENJIN PENGESAN MAKLUMAT LUKISAN DWG PDF
# ==============================================================================

def extract_dwg_info(pdf_source):
    """Mengekstrak maklumat lengkap daripada lukisan DWG PDF mengikut spesifikasi drafter."""
    if isinstance(pdf_source, (bytes, bytearray)):
        doc = pymupdf.open(stream=pdf_source, filetype="pdf")
        filename = "DWG_Drawing.pdf"
    elif isinstance(pdf_source, io.BytesIO):
        doc = pymupdf.open(stream=pdf_source.getvalue(), filetype="pdf")
        filename = "DWG_Drawing.pdf"
    else:
        if not os.path.exists(pdf_source):
            raise FileNotFoundError(f"Fail DWG PDF tidak ditemui: {pdf_source}")
        doc = pymupdf.open(pdf_source)
        filename = os.path.basename(pdf_source)

    all_text = ""
    pages_text = []
    for i, page in enumerate(doc):
        t = page.get_text("text")
        pages_text.append(t)
        all_text += f"\n--- Page {i+1} ---\n" + t

    p1_txt = pages_text[0] if pages_text else ""

    data = {
        "source_type": "dwg",
        "source_name": filename,
        "total_pages": len(doc),
        "client_name": "",
        "full_address": "",
        "short_address": "",
        "state": "Pulau Pinang",
        "latitude": None,
        "longitude": None,
        "coordinates_str": "",
        "system_size_kwp": 0.0,
        "system_size_kwac": 0.0,
        "pv_module": {
            "model": "ASTRO-N7 CHSM66RN-630W",
            "manufacturer": "Astronergy",
            "pan_file": "CHSM66RN(DG)F-BH-630.PAN",
            "unit_power_w": 630,
            "panel_type": "BIFACIAL",
            "total_panels": 0
        },
        "inverter": {
            "model": "HIS-5.0L-G3",
            "manufacturer": "Hoymiles",
            "unit_power_kw": 5.0,
            "total_units": 1,
            "model_series": "HIS",
            "phase": ""
        },
        "battery": {
            "manufacturer": "Hoymiles",
            "model": "LB-16D-G3",
            "units": 0,
            "voltage_v": 51.0,
            "capacity_ah": 314.0,
            "nominal_energy_kwh": 16.1
        },
        "strings": [],
        "orientations": [],
        "glob_hor_kwh_m2": 0,
        "used_energy_kwh": 0,
        "weather_data_variance": 3.8
    }

    # 1. CLIENT NAME & ADDRESS (Title Block)
    client_match = re.search(r"REV:\s*\n\s*\d{2}/\d{2}/\d{4}\s*\n\s*\d+\s*\n([^\n,]+),\s*\n(.*?)(?=\n(?:PV LAYOUT|MOUNTING|SINGLE LINE|INVERTER|SCALE|REV:|\d{2}\b))", all_text, re.DOTALL)
    if not client_match:
        client_match = re.search(r"\bFOR\s*\n([^\n,]+),\s*\n(.*?)(?=\n(?:PV LAYOUT|MOUNTING|SINGLE LINE|INVERTER|SCALE|REV:|\d{2}\b))", all_text, re.DOTALL)

    if client_match:
        data["client_name"] = client_match.group(1).strip()
        raw_addr = client_match.group(2).replace("\n", " ").strip()
        raw_addr = re.sub(r"\s+", " ", raw_addr)
        data["full_address"] = raw_addr

        # Short address: Ambil tulisan sebelum poskod / perkataan TAMAN / BANDAR
        short_candidate = re.split(r'(?i)\bBANDAR\b|\bTAMAN\b|\b\d{5}\b', raw_addr)[0].strip().rstrip(',')
        data["short_address"] = short_candidate if short_candidate else raw_addr

        # Kenalpasti Negeri & Weather variance
        addr_upper = raw_addr.upper()
        for st_key, st_name, w_rate in STATE_WEATHER_RATES:
            if st_key in addr_upper:
                data["state"] = st_name
                data["weather_data_variance"] = w_rate
                break

    # 2. KOORDINAT (GPS)
    coord_patterns = [
        r"COORDINATES?\s*[:\-]?\s*([0-9]+\.[0-9]+)\s*,\s*([0-9]+\.[0-9]+)",
        r"LOCATION\s*[:\-]?\s*([0-9]+\.[0-9]+)\s*,\s*([0-9]+\.[0-9]+)",
        r"([0-9]+\.[0-9]{4,})\s*,\s*([0-9]+\.[0-9]{4,})",
        r"([0-9]+\.[0-9]{4,})\s*\n\s*([0-9]+\.[0-9]{4,})"
    ]
    for cp in coord_patterns:
        m = re.search(cp, all_text)
        if m:
            c1, c2 = float(m.group(1)), float(m.group(2))
            if 1.0 <= c1 <= 7.5 and 99.0 <= c2 <= 120.0:
                data["latitude"], data["longitude"] = c1, c2
            elif 1.0 <= c2 <= 7.5 and 99.0 <= c1 <= 120.0:
                data["latitude"], data["longitude"] = c2, c1
            if data["latitude"] is not None:
                data["coordinates_str"] = f"{data['latitude']}, {data['longitude']}"
                break

    # 3. ROOF FALLS & ORIENTATIONS (Muka surat 1 - PV Layout)
    fall_matches = re.findall(r"FALL\s*\n?\s*(\d+)", p1_txt, re.IGNORECASE)
    fall_tilts = [int(f) for f in fall_matches]

    if len(fall_tilts) == 0:
        data["orientations"] = [
            {"id": 1, "str_val": "20/180°"},
            {"id": 2, "str_val": "0/0"},
            {"id": 3, "str_val": "0/0"},
            {"id": 4, "str_val": "0/0"}
        ]
    elif len(fall_tilts) == 1:
        t = fall_tilts[0]
        data["orientations"] = [
            {"id": 1, "str_val": f"{t}/180°"},
            {"id": 2, "str_val": "0/0"},
            {"id": 3, "str_val": "0/0"},
            {"id": 4, "str_val": "0/0"}
        ]
    elif len(fall_tilts) == 2:
        t1, t2 = fall_tilts[0], fall_tilts[1]
        data["orientations"] = [
            {"id": 1, "str_val": f"{t1}/103°"},
            {"id": 2, "str_val": f"{t2}/-167°"},
            {"id": 3, "str_val": "0/0"},
            {"id": 4, "str_val": "0/0"}
        ]
    elif len(fall_tilts) == 3:
        t1, t2, t3 = fall_tilts[0], fall_tilts[1], fall_tilts[2]
        data["orientations"] = [
            {"id": 1, "str_val": f"{t1}/103°"},
            {"id": 2, "str_val": f"{t2}/-167°"},
            {"id": 3, "str_val": f"{t3}/13°"},
            {"id": 4, "str_val": "0/0"}
        ]
    else:
        t1, t2, t3, t4 = fall_tilts[0], fall_tilts[1], fall_tilts[2], fall_tilts[3]
        data["orientations"] = [
            {"id": 1, "str_val": f"{t1}/103°"},
            {"id": 2, "str_val": f"{t2}/-167°"},
            {"id": 3, "str_val": f"{t3}/13°"},
            {"id": 4, "str_val": f"{t4}/-77°"}
        ]

    # 4. SYSTEM SIZE (kWp & kWac)
    kwp_m = re.search(r"(\d+(?:\.\d+)?)\s*k(?:Wp|WP)\b", all_text)
    if kwp_m:
        data["system_size_kwp"] = float(kwp_m.group(1))

    kwac_m = re.search(r"(\d+(?:\.\d+)?)\s*k(?:Wac|WAC)\b", all_text)
    if kwac_m:
        data["system_size_kwac"] = float(kwac_m.group(1))

    # 5. SPESIFIKASI MODUL PV (Astronergy vs JA Solar 630/635/640)
    has_630 = bool(re.search(r"\b630\b|CHSM66RN|ASTRO", all_text, re.IGNORECASE))
    has_635 = bool(re.search(r"\b635\b|JAM72D42-635", all_text, re.IGNORECASE))
    has_640 = bool(re.search(r"\b640\b|JAM72D42-640", all_text, re.IGNORECASE))
    has_ja = bool(re.search(r"\bJA\b|JASOLAR", all_text, re.IGNORECASE))

    if has_635 or (has_ja and not has_640 and not has_630):
        data["pv_module"]["manufacturer"] = "JA Solar"
        data["pv_module"]["model"] = "JAM72D42-635/LB"
        data["pv_module"]["pan_file"] = "JAM72D42-635/LB.PAN"
        data["pv_module"]["unit_power_w"] = 635
    elif has_640 or (has_ja and has_640):
        data["pv_module"]["manufacturer"] = "JA Solar"
        data["pv_module"]["model"] = "JAM72D42-640/LB"
        data["pv_module"]["pan_file"] = "JAM72D42-640/LB.PAN"
        data["pv_module"]["unit_power_w"] = 640
    else:
        data["pv_module"]["manufacturer"] = "Astronergy"
        data["pv_module"]["model"] = "ASTRO-N7 CHSM66RN-630W"
        data["pv_module"]["pan_file"] = "CHSM66RN(DG)F-BH-630.PAN"
        data["pv_module"]["unit_power_w"] = 630

    # Total panel count
    panel_count_m = re.search(r"(\d+)\s*PCS\s*PV\s*MODULE", all_text, re.IGNORECASE)
    if not panel_count_m:
        panel_count_m = re.search(r"TOTAL\s*PANELS?\s*[:\-]?\s*(\\d+)", all_text, re.IGNORECASE)
    if panel_count_m:
        data["pv_module"]["total_panels"] = int(panel_count_m.group(1))
        if data["system_size_kwp"] == 0.0:
            data["system_size_kwp"] = round((data["pv_module"]["total_panels"] * data["pv_module"]["unit_power_w"]) / 1000, 2)

    # 6. INVERTER MODEL (KOD SAHAJA TANPA PERKATAAN HOYMILES)
    inv_m = re.search(r"SYSTEM\s*INFORMATION.*?\bINVERTER\s*\n\s*(\d+(?:\\.\\d+)?)\s*x\s*([^\n]+)", all_text, re.DOTALL)
    if not inv_m:
        inv_m = re.search(r"\bINVERTER\s*\n\s*(\d+(?:\\.\\d+)?)\s*x\s*([^\n]+)", all_text)

    if inv_m:
        data["inverter"]["total_units"] = int(float(inv_m.group(1)))
        raw_inv = inv_m.group(2).strip()
        clean_inv = re.sub(r'(?i)\bHOYMILES\b|\bHUAWEI\b|\bSUNGROW\b', '', raw_inv).strip()
        data["inverter"]["model"] = clean_inv
        if "HIT" in clean_inv.upper():
            data["inverter"]["model_series"] = "HIT"
        elif "HIS" in clean_inv.upper():
            data["inverter"]["model_series"] = "HIS"

        p_m = re.search(r"(\d+(?:\.\d+)?)\s*L-G", clean_inv, re.IGNORECASE)
        if p_m:
            data["inverter"]["unit_power_kw"] = float(p_m.group(1))
        elif data["system_size_kwac"] > 0:
            data["inverter"]["unit_power_kw"] = data["system_size_kwac"]

    if data["inverter"]["unit_power_kw"] == 0.0 and data["system_size_kwac"] > 0:
        data["inverter"]["unit_power_kw"] = data["system_size_kwac"]

    # 7. BATTERY (Pengecaman Pintar: 01, 1.0, 1, 2, tiada bateri, 0.0, 0, -, N/A)
    # Pola kuantiti bateri (cth: '1.0 x HOYMILES LB-16D-G3', '01 x LB-16D', '1 x BATTERY', '0.0 x ...')
    bat_patterns = [
        r'(\d+(?:\.\d+)?)\s*x\s*(?:HOYMILES\s*)?(LB[^\n,]+|BATTERY[^\n,]*)',
        r'BATTERY\s*[:\-]?\s*(\d+(?:\.\d+)?)\s*(?:x|PCS|UNIT|UNITS)?\s*([^\n]*)',
        r'(\d+(?:\.\d+)?)\s*(?:x|PCS|UNIT|UNITS)\s*(?:HOYMILES\s*)?(LB[^\n,]*)',
        r'(\d+(?:\.\d+)?)\s*x\s*(?:HOYMILES\s*)?(?:BATTERY|ENERGY\s*STORAGE)',
    ]
    
    matched_bat_qty = None
    matched_bat_model = None
    for bp in bat_patterns:
        bm = re.search(bp, all_text, re.IGNORECASE)
        if bm:
            try:
                matched_bat_qty = float(bm.group(1))
                if len(bm.groups()) >= 2 and bm.group(2):
                    matched_bat_model = bm.group(2).strip()
                break
            except Exception:
                pass

    if matched_bat_qty is not None:
        if matched_bat_qty <= 0.0:
            data["battery"]["units"] = 0
            data["battery"]["model"] = ""
        else:
            data["battery"]["units"] = int(round(matched_bat_qty))
            if matched_bat_model:
                data["battery"]["model"] = re.sub(r'(?i)\bHOYMILES\b', '', matched_bat_model).strip()
    elif re.search(r'BATTERY\s*[:\-]?\s*(?:0(?:\.0+)?|-|N/?A|NIL|NONE|TIADA)\b', all_text, re.IGNORECASE):
        # Explicitly mention zero or no battery
        data["battery"]["units"] = 0
        data["battery"]["model"] = ""
    else:
        # Check if LB-16D / LB- is mentioned anywhere in the drawing
        lb_match = re.search(r'(?:HOYMILES\s*)?(LB-[0-9A-Z\-]+)', all_text, re.IGNORECASE)
        if lb_match:
            data["battery"]["units"] = 1
            data["battery"]["model"] = lb_match.group(1).strip()
        else:
            # Tiada bateri langsung dalam drawing
            data["battery"]["units"] = 0
            data["battery"]["model"] = ""

    # 8. STRINGS CONFIGURATION
    str_block = re.search(r"STRING\s*\n\s*CONFIGURATION\s*\n((?:\d+\s*PCS\s*X\s*\d+\s*STG\s*\n?)+)", all_text, re.IGNORECASE)
    if str_block:
        lines_s = [l.strip() for l in str_block.group(1).split("\n") if l.strip()]
        for idx, line in enumerate(lines_s):
            sm = re.search(r"(\d+)\s*PCS\s*X\s*(\d+)\s*STG", line, re.IGNORECASE)
            if sm:
                pcs = int(sm.group(1))
                stg = int(sm.group(2))
                data["strings"].append({
                    "array_id": idx + 1,
                    "panels_per_string": pcs,
                    "strings_count": stg,
                    "total_panels": pcs * stg
                })

    if not data["strings"]:
        inv_str_matches = re.findall(r"INV\s*-\s*STR\s*\(\s*\)\s*\n\s*1\s*\n\s*(\d+)\s+(\d+)", all_text)
        for idx, m in enumerate(inv_str_matches):
            data["strings"].append({
                "array_id": idx + 1,
                "panels_per_string": int(m[1]),
                "strings_count": 1,
                "total_panels": int(m[1])
            })

    return data

# ==============================================================================
# 3. ENJIN PENGESAN MAKLUMAT FAIL EXCEL INPUT (PAGE 1 HINGGA PAGE 7)
# ==============================================================================

def extract_excel_info(excel_source):
    """
    Mengekstrak maklumat lengkap projek daripada fail Excel (Template.xlsx yang disemak),
    termasuk PAGE 1, PAGE2, PAGE 3, PAGE 4, PAGE 5, PAGE 6, dan PAGE 7.
    """
    if isinstance(excel_source, (bytes, bytearray)):
        wb = openpyxl.load_workbook(io.BytesIO(excel_source), data_only=True)
        filename = "Input_Excel.xlsx"
    elif isinstance(excel_source, io.BytesIO):
        wb = openpyxl.load_workbook(excel_source, data_only=True)
        filename = "Input_Excel.xlsx"
    else:
        if not os.path.exists(excel_source):
            raise FileNotFoundError(f"Fail Excel tidak ditemui: {excel_source}")
        wb = openpyxl.load_workbook(excel_source, data_only=True)
        filename = os.path.basename(excel_source)

    ex_data = {
        "source_type": "excel",
        "source_name": filename,
        "client_name": "",
        "short_address": "",
        "full_address": "",
        "state": "Pulau Pinang",
        "latitude": None,
        "longitude": None,
        "system_size_kwp": 0.0,
        "system_size_kwac": 0.0,
        "glob_hor_kwh_m2": 0,
        "used_energy_kwh": 0,
        "produced_energy_kwh": 0.0,
        "specific_prod_kwh_kwp": 0.0,
        "perf_ratio_pr": 0.0,
        "solar_fraction_sf": 0.0,
        "weather_data_variance": None,
        "p50_variability_kwh": 0,
        "p50_kwh": 0,
        "p90_kwh": 0,
        "p75_kwh": 0,
        "pv_module": {
            "manufacturer": "Astronergy",
            "model": "ASTRO-N7 CHSM66RN-630W",
            "pan_file": "CHSM66RN(DG)F-BH-630.PAN",
            "unit_power_w": 630,
            "total_panels": 0
        },
        "inverter": {
            "manufacturer": "Hoymiles",
            "model": "HIS-5.0L-G3",
            "unit_power_kw": 5.0,
            "total_units": 1,
            "model_series": "HIS"
        },
        "battery": {
            "manufacturer": "Hoymiles",
            "model": "LB-16D-G3",
            "units": 0,
            "voltage_v": 51.0,
            "capacity_ah": 314.0,
            "nominal_energy_kwh": 16.1
        },
        "orientations": [],
        "strings": [],
        "arrays": []
    }

    # 1. PAGE 1
    if "PAGE 1" in wb.sheetnames:
        ws1 = wb["PAGE 1"]
        ex_data["client_name"] = str(ws1["B1"].value or "").strip()
        try:
            if ws1["B2"].value is not None:
                ex_data["latitude"] = float(ws1["B2"].value)
        except Exception:
            pass
        try:
            if ws1["C2"].value is not None:
                ex_data["longitude"] = float(ws1["C2"].value)
        except Exception:
            pass

        raw_s = str(ws1["B3"].value or "").strip()
        ex_data["short_address"] = re.split(r'(?i)\bBANDAR\b|\bTAMAN\b|\b\d{5}\b', raw_s)[0].strip().rstrip(',')
        ex_data["full_address"] = raw_s

        try:
            if ws1["B5"].value is not None:
                ex_data["glob_hor_kwh_m2"] = int(float(ws1["B5"].value))
        except Exception:
            pass
        try:
            if ws1["B6"].value is not None:
                ex_data["used_energy_kwh"] = int(float(ws1["B6"].value))
        except Exception:
            pass

    if ex_data["short_address"]:
        addr_up = ex_data["short_address"].upper()
        for st_key, st_name, w_rate in STATE_WEATHER_RATES:
            if st_key in addr_up:
                ex_data["state"] = st_name
                ex_data["weather_data_variance"] = w_rate
                break

    # 2. PAGE2
    if "PAGE2" in wb.sheetnames:
        ws2 = wb["PAGE2"]
        for idx, c in enumerate(["B9", "D9", "F9", "I9"]):
            val = str(ws2[c].value or "").strip()
            if val and val not in ["0/0", "0/0°", "0", "None", ""]:
                ex_data["orientations"].append({"id": idx + 1, "tilt_azimuth": val, "str_val": val})

        try:
            if ws2["B13"].value is not None:
                ex_data["pv_module"]["total_panels"] = int(float(ws2["B13"].value))
        except Exception:
            pass
        try:
            if ws2["B14"].value is not None:
                ex_data["system_size_kwp"] = float(ws2["B14"].value)
        except Exception:
            pass
        try:
            if ws2["D13"].value is not None:
                ex_data["inverter"]["total_units"] = int(float(ws2["D13"].value))
        except Exception:
            pass
        try:
            if ws2["D14"].value is not None:
                ex_data["inverter"]["unit_power_kw"] = float(ws2["D14"].value)
                ex_data["system_size_kwac"] = ex_data["inverter"]["unit_power_kw"]
        except Exception:
            pass
        try:
            if ws2["D15"].value is not None:
                ex_data["inverter"]["pnom_ratio"] = float(ws2["D15"].value)
        except Exception:
            pass
        try:
            if ws2["F13"].value is not None:
                ex_data["battery"]["units"] = int(float(ws2["F13"].value))
        except Exception:
            pass
        try:
            if ws2["F14"].value is not None:
                ex_data["battery"]["voltage_v"] = float(ws2["F14"].value)
        except Exception:
            pass
        try:
            if ws2["F15"].value is not None:
                ex_data["battery"]["capacity_ah"] = float(ws2["F15"].value)
        except Exception:
            pass

        # Results summary
        try:
            if ws2["B18"].value is not None:
                ex_data["produced_energy_kwh"] = float(ws2["B18"].value)
        except Exception:
            pass
        try:
            if ws2["D18"].value is not None:
                ex_data["specific_prod_kwh_kwp"] = float(ws2["D18"].value)
        except Exception:
            pass
        try:
            if ws2["F18"].value is not None:
                ex_data["perf_ratio_pr"] = float(ws2["F18"].value)
        except Exception:
            pass
        try:
            if ws2["B19"].value is not None:
                ex_data["used_energy_kwh"] = float(ws2["B19"].value)
        except Exception:
            pass
        try:
            if ws2["F19"].value is not None:
                ex_data["solar_fraction_sf"] = float(ws2["F19"].value)
        except Exception:
            pass

    # 3. PAGE 3
    if "PAGE 3" in wb.sheetnames:
        ws3 = wb["PAGE 3"]
        # Extract Monthly Self Consumption
        monthly_cells = ["B6", "B7", "B8", "B9", "B10", "D6", "D7", "D8", "D9", "D10", "F6", "F7", "F8"]
        ex_data["monthly_consumption"] = []
        for c_ref in monthly_cells:
            try:
                c_val = ws3[c_ref].value
                if c_val is not None:
                    ex_data["monthly_consumption"].append(float(c_val))
                else:
                    ex_data["monthly_consumption"].append(0.0)
            except Exception:
                ex_data["monthly_consumption"].append(0.0)
        if ws3["B14"].value:
            ex_data["pv_module"]["manufacturer"] = str(ws3["B14"].value).strip()
        if ws3["B15"].value:
            ex_data["pv_module"]["model"] = str(ws3["B15"].value).strip()
        if ws3["D15"].value:
            raw_inv = str(ws3["D15"].value).strip()
            ex_data["inverter"]["model"] = re.sub(r'(?i)\bHOYMILES\b|\bHUAWEI\b|\bSUNGROW\b', '', raw_inv).strip()
        if ws3["B16"].value:
            ex_data["pv_module"]["pan_file"] = str(ws3["B16"].value).strip()
        try:
            if ws3["B17"].value is not None:
                ex_data["pv_module"]["unit_power_w"] = int(float(ws3["B17"].value))
        except Exception:
            pass
        try:
            if ws3["D17"].value is not None:
                ex_data["inverter"]["unit_power_kw"] = float(ws3["D17"].value)
        except Exception:
            pass

        if "HIT" in ex_data["inverter"]["model"].upper():
            ex_data["inverter"]["model_series"] = "HIT"
        elif "HIS" in ex_data["inverter"]["model"].upper():
            ex_data["inverter"]["model_series"] = "HIS"

        # Arrays 1 to 4
        row_offsets = [19, 29, 39, 49]
        is_hit = ex_data["inverter"]["model_series"] == "HIT"
        for arr_idx, r in enumerate(row_offsets):
            mod_count = ws3.cell(r+3, 2).value
            if mod_count and float(mod_count) > 0:
                p_cnt = int(float(mod_count))
                ex_data["strings"].append({"array_id": arr_idx + 1, "total_panels": p_cnt})

                op_volt = ws3.cell(r+6, 5 if is_hit else 4).value
                pnom_r = ws3.cell(r+7, 5 if is_hit else 4).value
                inv_pct = ws3.cell(r+3, 4).value
                inv_unit = ws3.cell(r+3, 5).value
                tot_pwr = ws3.cell(r+4, 4).value

                arr_info = {
                    "array_id": arr_idx + 1,
                    "orientation": ws3.cell(r+1, 2).value or (arr_idx + 1),
                    "tilt_azimuth": str(ws3.cell(r+2, 2).value or ""),
                    "modules": p_cnt,
                    "stc_wp": float(ws3.cell(r+4, 2).value or (p_cnt * ex_data["pv_module"]["unit_power_w"])),
                    "modules_series": int(float(ws3.cell(r+5, 2).value or p_cnt)),
                    "pmpp": float(ws3.cell(r+7, 2).value or 0.0),
                    "umpp": float(ws3.cell(r+8, 2).value or 0.0),
                    "impp": float(ws3.cell(r+9, 2).value or 0.0),
                    "mppt_pct": int(float(inv_pct)) if inv_pct is not None else None,
                    "inv_unit_fraction": float(inv_unit) if inv_unit is not None else None,
                    "inverter_power_kwac": float(tot_pwr) if tot_pwr is not None else None,
                    "operating_voltage": str(op_volt or ("150-900" if is_hit else "80-500")),
                    "pnom_ratio": float(pnom_r) if pnom_r is not None else None
                }
                ex_data["arrays"].append(arr_info)

        # 3b. PAGE 4: Grid Storage, DC & AC Wiring Losses
        if "PAGE 4" in wb.sheetnames:
            ws4 = wb["PAGE 4"]
            try:
                if ws4["D3"].value is not None:
                    ex_data["battery"]["max_charge_power"] = float(ws4["D3"].value)
            except Exception:
                pass
            try:
                if ws4["D5"].value is not None:
                    ex_data["battery"]["max_discharge_power"] = float(ws4["D5"].value)
            except Exception:
                pass
            try:
                if ws4["B5"].value is not None:
                    ex_data["battery"]["capacity_ah"] = float(ws4["B5"].value)
            except Exception:
                pass
            try:
                if ws4["B3"].value is not None:
                    ex_data["battery"]["units"] = int(float(ws4["B3"].value))
            except Exception:
                pass
            try:
                if ws4["B4"].value is not None:
                    ex_data["battery"]["voltage_v"] = float(ws4["B4"].value)
            except Exception:
                pass

            # DC wiring losses calculation
            l1 = float(ws4["B10"].value or 100) if ws4["B10"].value is not None else 100.0
            l2 = float(ws4["D10"].value or 80) if ws4["D10"].value is not None else 80.0
            l3 = float(ws4["B14"].value or 50) if ws4["B14"].value is not None else 50.0
            l4 = float(ws4["D14"].value or 80) if ws4["D14"].value is not None else 80.0

            u_w = float(ex_data["pv_module"]["unit_power_w"] or 630.0)
            arrs = ex_data.get("arrays", [])
            wp_b23 = arrs[0]["stc_wp"] if len(arrs) > 0 and arrs[0].get("stc_wp") else (6 * u_w)
            wp_b33 = arrs[1]["stc_wp"] if len(arrs) > 1 and arrs[1].get("stc_wp") else (8 * u_w)
            wp_b43 = arrs[2]["stc_wp"] if len(arrs) > 2 and arrs[2].get("stc_wp") else (5 * u_w)
            wp_b53 = arrs[3]["stc_wp"] if len(arrs) > 3 and arrs[3].get("stc_wp") else (5 * u_w)

            r1_c = int(round(0.0175 * l1 / 6.0 * 1000.0))
            r2_c = int(round(0.0175 * l2 / 6.0 * 1000.0))
            r3_c = int(round(0.0175 * l3 / 6.0 * 1000.0))
            r4_c = int(round(0.0175 * l4 / 6.0 * 1000.0))

            # In Excel template: D12 references B43, B16 references B33
            loss1_c = round((256.0 * (0.0175 * l1 / 6.0) / wp_b23) * 100.0, 2) if wp_b23 > 0 else 1.98
            loss2_c = round((256.0 * (0.0175 * l2 / 6.0) / wp_b43) * 100.0, 2) if wp_b43 > 0 else 1.90
            loss3_c = round((256.0 * (0.0175 * l3 / 6.0) / wp_b33) * 100.0, 2) if wp_b33 > 0 else 0.74
            loss4_c = round((256.0 * (0.0175 * l4 / 6.0) / wp_b53) * 100.0, 2) if wp_b53 > 0 else 1.90

            def _get_f(val, fallback):
                try:
                    if val is not None:
                        return float(val)
                except Exception:
                    pass
                return fallback

            r1 = int(round(_get_f(ws4["B11"].value, r1_c)))
            l1_val = _get_f(ws4["B12"].value, loss1_c)

            r2 = int(round(_get_f(ws4["D11"].value, r2_c)))
            l2_val = _get_f(ws4["D12"].value, loss2_c)

            r3 = int(round(_get_f(ws4["B15"].value, r3_c)))
            l3_val = _get_f(ws4["B16"].value, loss3_c)

            r4 = int(round(_get_f(ws4["D15"].value, r4_c)))
            l4_val = _get_f(ws4["D16"].value, loss4_c)

            arr_cnt = len(arrs) if len(arrs) > 0 else 2
            if arr_cnt == 2:
                act_losses = [l1_val, l2_val]
            elif arr_cnt == 3:
                act_losses = [l1_val, l2_val, l3_val]
            else:
                act_losses = [l1_val, l2_val, l3_val, l4_val]

            gl_c = round(sum(act_losses) / len(act_losses), 2) if act_losses else 1.63
            gl_val = _get_f(ws4["B9"].value, gl_c)

            ex_data["dc_losses"] = {
                "global_loss": gl_val,
                "arr1_res": r1,
                "arr1_loss": l1_val,
                "arr2_res": r2,
                "arr2_loss": l2_val,
                "arr3_res": r3,
                "arr3_loss": l3_val,
                "arr4_res": r4,
                "arr4_loss": l4_val,
            }

            # AC wiring losses
            is_hit_model = "HIT" in str(ex_data["inverter"]["model"]).upper()
            inv_volt_val = ws4["C19"].value if is_hit_model else ws4["B19"].value
            wire_sec_val = ws4["C21"].value if is_hit_model else ws4["B21"].value
            wire_len_val = ws4["C22"].value if is_hit_model else ws4["B22"].value

            sec_str = str(wire_sec_val or ("3x3" if is_hit_model else "2x4")).strip()
            sec_str = sec_str.replace("x", " x ").replace("X", " x ")

            ex_data["ac_losses"] = {
                "inverter_voltage": f"{inv_volt_val or (400 if is_hit_model else 220)} Vac" + (" tri" if is_hit_model else " mono"),
                "wire_section": f"Copper {sec_str} mm²",
                "wires_length": f"{int(float(wire_len_val or (20 if is_hit_model else 10)))} m"
            }

    # 4. PAGE 5 (Main Results, PR & Balances Table)
    if "PAGE 5" in wb.sheetnames:
        ws5 = wb["PAGE 5"]
        try:
            if ws5["B3"].value is not None:
                ex_data["produced_energy_kwh"] = float(ws5["B3"].value)
        except Exception:
            pass
        try:
            if ws5["D3"].value is not None:
                ex_data["specific_prod_kwh_kwp"] = float(ws5["D3"].value)
        except Exception:
            pass
        try:
            if ws5["D4"].value is not None:
                ex_data["perf_ratio_pr"] = float(ws5["D4"].value)
        except Exception:
            pass
        try:
            if ws5["B4"].value is not None:
                ex_data["used_energy_kwh"] = float(ws5["B4"].value)
        except Exception:
            pass
        try:
            if ws5["D5"].value is not None:
                ex_data["solar_fraction_sf"] = float(ws5["D5"].value)
        except Exception:
            pass

        # Ekstrak Jadual Balances and main results (13 baris x 10 lajur)
        bal_table = []
        for r_idx in range(9, 22):
            r_vals = []
            for c_idx in range(2, 12):
                cell_v = ws5.cell(r_idx, c_idx).value
                if cell_v is not None:
                    try:
                        fv = float(cell_v)
                        if c_idx == 4: # T_Amb
                            str_v = f"{fv:.2f}" if r_idx == 21 else f"{fv:.1f}"
                        elif fv.is_integer():
                            str_v = str(int(fv))
                        else:
                            str_v = f"{fv:.1f}"
                    except Exception:
                        str_v = str(cell_v)
                else:
                    str_v = "0"
                r_vals.append(str_v)
            bal_table.append(r_vals)
        ex_data["balances_table"] = bal_table

    # 5. PAGE 6 (Loss Diagram)
    if "PAGE 6" in wb.sheetnames:
        ws6 = wb["PAGE 6"]
        loss_dict = {}
        try:
            if ws6["B3"].value is not None:
                loss_dict["glob_hor"] = float(ws6["B3"].value)
        except Exception:
            pass
        try:
            if ws6["B4"].value is not None:
                loss_dict["glob_inc"] = float(ws6["B4"].value)
        except Exception:
            pass
        try:
            if ws6["B5"].value is not None:
                loss_dict["pv_nom"] = float(ws6["B5"].value)
        except Exception:
            pass
        try:
            if ws6["B6"].value is not None:
                loss_dict["array_virt"] = float(ws6["B6"].value)
        except Exception:
            pass
        try:
            if ws6["B7"].value is not None:
                loss_dict["final_energy"] = float(ws6["B7"].value)
        except Exception:
            pass
        try:
            if ws6["B8"].value is not None:
                loss_dict["efr_grid"] = float(ws6["B8"].value)
        except Exception:
            pass
        try:
            if ws6["B9"].value is not None:
                loss_dict["battery_stored"] = float(ws6["B9"].value)
        except Exception:
            pass
        try:
            if ws6["B10"].value is not None:
                loss_dict["direct_use"] = float(ws6["B10"].value)
        except Exception:
            pass
        ex_data["loss_diagram"] = loss_dict

    # 6. PAGE 7 (P50 - P90)
    if "PAGE 7" in wb.sheetnames:
        ws7 = wb["PAGE 7"]
        try:
            if ws7["B2"].value is not None:
                ex_data["weather_data_variance"] = float(ws7["B2"].value)
        except Exception:
            pass
        try:
            if ws7["B6"].value is not None:
                ex_data["p50_variability_kwh"] = round(float(ws7["B6"].value))
        except Exception:
            pass
        try:
            if ws7["B7"].value is not None:
                ex_data["p50_kwh"] = round(float(ws7["B7"].value))
        except Exception:
            pass
        try:
            if ws7["B8"].value is not None:
                ex_data["p90_kwh"] = round(float(ws7["B8"].value))
        except Exception:
            pass
        try:
            if ws7["B9"].value is not None:
                ex_data["p75_kwh"] = round(float(ws7["B9"].value))
        except Exception:
            pass

    if ex_data["system_size_kwp"] == 0.0 and ex_data["pv_module"]["total_panels"] > 0 and ex_data["pv_module"]["unit_power_w"] > 0:
        ex_data["system_size_kwp"] = round((ex_data["pv_module"]["total_panels"] * ex_data["pv_module"]["unit_power_w"]) / 1000, 2)

    # Automatic calculation fallback if draft Excel has uncalculated formulas for PAGE 5/6/7
    if ex_data["produced_energy_kwh"] == 0.0 or not ex_data.get("balances_table") or len(ex_data.get("balances_table", [])) < 12:
        kwp = float(ex_data.get("system_size_kwp") or 8.82)
        yr_in = float(ex_data.get("glob_hor_kwh_m2") or 1715)
        load = float(ex_data.get("used_energy_kwh") or 10350)
        if load == 0: load = 10350.0
        if yr_in == 0: yr_in = 1715.0

        m_pcts = [0.08, 0.07, 0.06, 0.09, 0.08, 0.09, 0.08, 0.08, 0.09, 0.08, 0.06, 0.07]
        diff_r = [0.55, 0.52, 0.48, 0.50, 0.53, 0.49, 0.45, 0.46, 0.50, 0.55, 0.62, 0.60]
        t_amb = [27.5, 27.8, 28.2, 28.5, 28.8, 28.9, 28.7, 28.6, 28.3, 28.0, 27.7, 27.5]
        pf_arr = [0.82, 0.81, 0.78, 0.82, 0.83, 0.82, 0.81, 0.79, 0.82, 0.83, 0.79, 0.84]
        sp_user = [0.08, 0.075, 0.08, 0.085, 0.09, 0.095, 0.09, 0.085, 0.08, 0.08, 0.075, 0.085]

        # Calculate dynamic TF average from actual orientation panels
        arr_mods = [a.get("modules", 0) for a in ex_data.get("arrays", [])]
        while len(arr_mods) < 4:
            arr_mods.append(0)
        tot_mods_cnt = sum(arr_mods) or int(ex_data["pv_module"]["total_panels"] or 14)
        tfs = [0.97, 0.95, 0.97, 0.97]
        tf_avg = sum(arr_mods[i] * tfs[i] for i in range(4)) / tot_mods_cnt if tot_mods_cnt > 0 else 0.9567

        tot_gh, tot_df, tot_gi, tot_ge, tot_ea, tot_eu, tot_es, tot_eg, tot_ef = 0, 0, 0, 0, 0, 0, 0, 0, 0
        table = []

        for i in range(12):
            gh = round(m_pcts[i] * yr_in, 1)
            df = round(gh * diff_r[i], 1)
            gi = round(gh * tf_avg, 1)
            ge = round(gi * 0.95, 1)
            ea = round(ge * kwp * pf_arr[i], 1)
            eu = round(sp_user[i] * load, 1)
            es = round(min(ea, eu), 1)
            eg = round(max(ea - es, 0), 1)
            ef = round(max(eu - es, 0), 1)

            tot_gh += gh; tot_df += df; tot_gi += gi; tot_ge += ge; tot_ea += ea; tot_eu += eu; tot_es += es; tot_eg += eg; tot_ef += ef
            table.append([f"{gh:.1f}", f"{df:.1f}", f"{t_amb[i]:.1f}", f"{gi:.1f}", f"{ge:.1f}", f"{ea:.1f}", f"{eu:.1f}", f"{es:.1f}", f"{eg:.1f}", f"{ef:.1f}"])

        table.append([f"{tot_gh:.1f}", f"{tot_df:.1f}", "28.21", f"{tot_gi:.1f}", f"{tot_ge:.1f}", f"{tot_ea:.1f}", f"{tot_eu:.1f}", f"{tot_es:.1f}", f"{tot_eg:.1f}", f"{tot_ef:.1f}"])

        ex_data["produced_energy_kwh"] = round(tot_ea, 1)
        ex_data["specific_prod_kwh_kwp"] = round(tot_ea / kwp, 1)
        ex_data["perf_ratio_pr"] = 81.45
        ex_data["used_energy_kwh"] = round(tot_eu, 1)
        ex_data["solar_fraction_sf"] = round((tot_es / tot_eu) * 100, 2) if tot_eu > 0 else 100.0
        ex_data["balances_table"] = table

        if sum(ex_data.get("monthly_consumption", [])) == 0.0:
            m_calc = [round(sp_user[i] * tot_eu, 1) for i in range(12)]
            m_calc.append(round(sum(m_calc), 1))
            ex_data["monthly_consumption"] = m_calc

        has_bat_flag = True if (ex_data.get("has_battery") or (ex_data.get("battery", {}).get("units", 0) > 0)) else False
        virt = round(tot_ea * (1 - 0.085), 1)
        avail = round(virt * (1 - 0.006) * (1 - 0.0264), 1)
        bat_stored = round(tot_es * 0.8, 1) if has_bat_flag else 0.0
        direct_use = round(tot_eg, 1) if has_bat_flag else round(avail, 1)

        ex_data["loss_diagram"] = {
            "glob_hor": tot_gh,
            "glob_inc": tot_gi,
            "pv_nom": tot_ea,
            "array_virt": virt,
            "final_energy": avail,
            "efr_grid": tot_ef,
            "battery_stored": bat_stored,
            "direct_use": direct_use
        }

        v_n = float(ex_data.get("weather_data_variance") or 4.0)
        cv_n = float(ex_data.get("custom_variability") or 5.7)
        sys_unc = math.sqrt(0.01**2 + 0.005**2 + 0.01**2 + 0.01**2 + (v_n / 100.0)**2)
        glob_var = math.sqrt((cv_n / 100.0)**2 + sys_unc**2)
        p50_kwh = round(tot_ea)
        var_kwh = round(p50_kwh * glob_var)
        p90_kwh = round(p50_kwh - (1.282 * glob_var * p50_kwh))
        p75_kwh = round(p50_kwh - (0.674 * glob_var * p50_kwh))

        ex_data["p50_p90"] = {
            "variance": v_n,
            "custom_variability": cv_n,
            "variability": var_kwh,
            "p50": p50_kwh,
            "p90": p90_kwh,
            "p75": p75_kwh
        }
        ex_data["p50_variability_kwh"] = var_kwh
        ex_data["p50_kwh"] = p50_kwh
        ex_data["p90_kwh"] = p90_kwh
        ex_data["p75_kwh"] = p75_kwh

    return ex_data

# ==============================================================================
# 4. GABUNGKAN DATA PROJEK (HARMONISASI)
# ==============================================================================

def combine_project_data(dwg_data=None, pv_data=None, excel_data=None):
    """Menggabungkan sumber data daripada DWG, PVsyst, atau Excel ke dalam satu kamus projek."""
    base = {
        "client_name": "",
        "short_address": "",
        "full_address": "",
        "state": "Pulau Pinang",
        "latitude": None,
        "longitude": None,
        "system_size_kwp": 0.0,
        "system_size_kwac": 0.0,
        "glob_hor_kwh_m2": 0,
        "used_energy_kwh": 0,
        "produced_energy_kwh": 0.0,
        "specific_prod_kwh_kwp": 0.0,
        "perf_ratio_pr": 0.0,
        "solar_fraction_sf": 0.0,
        "weather_data_variance": None,
        "p50_variability_kwh": 0,
        "p50_kwh": 0,
        "p90_kwh": 0,
        "p75_kwh": 0,
        "pv_module": {
            "manufacturer": "Astronergy",
            "model": "ASTRO-N7 CHSM66RN-630W",
            "pan_file": "CHSM66RN(DG)F-BH-630.PAN",
            "unit_power_w": 630,
            "total_panels": 0
        },
        "inverter": {
            "manufacturer": "Hoymiles",
            "model": "HIS-5.0L-G3",
            "unit_power_kw": 5.0,
            "total_units": 1,
            "model_series": "HIS"
        },
        "battery": {
            "manufacturer": "Hoymiles",
            "model": "LB-16D-G3",
            "units": 0,
            "voltage_v": 51.0,
            "capacity_ah": 314.0,
            "nominal_energy_kwh": 16.1
        },
        "orientations": [],
        "strings": [],
        "arrays": []
    }

    if dwg_data:
        for k in ["client_name", "short_address", "full_address", "state", "latitude", "longitude", "system_size_kwp", "system_size_kwac", "weather_data_variance"]:
            if dwg_data.get(k) is not None:
                base[k] = dwg_data[k]
        if dwg_data.get("orientations"):
            base["orientations"] = dwg_data["orientations"]
        if dwg_data.get("pv_module"):
            base["pv_module"].update({k: v for k, v in dwg_data["pv_module"].items() if v is not None and v != ""})
        if dwg_data.get("inverter"):
            base["inverter"].update({k: v for k, v in dwg_data["inverter"].items() if v is not None and v != ""})
        if dwg_data.get("battery"):
            base["battery"].update({k: v for k, v in dwg_data["battery"].items() if v is not None})
        if dwg_data.get("strings"):
            base["strings"] = dwg_data["strings"]

    if excel_data:
        for k in ["client_name", "short_address", "full_address", "state", "latitude", "longitude", "system_size_kwp", "system_size_kwac", "glob_hor_kwh_m2", "used_energy_kwh", "produced_energy_kwh", "specific_prod_kwh_kwp", "perf_ratio_pr", "solar_fraction_sf", "weather_data_variance", "p50_variability_kwh", "p50_kwh", "p90_kwh", "p75_kwh"]:
            if excel_data.get(k) is not None and excel_data.get(k) != 0:
                base[k] = excel_data[k]
        if excel_data.get("orientations"):
            base["orientations"] = excel_data["orientations"]
        if excel_data.get("pv_module"):
            base["pv_module"].update({k: v for k, v in excel_data["pv_module"].items() if v is not None and v != ""})
        if excel_data.get("inverter"):
            base["inverter"].update({k: v for k, v in excel_data["inverter"].items() if v is not None and v != ""})
        if excel_data.get("battery"):
            base["battery"].update({k: v for k, v in excel_data["battery"].items() if v is not None})
        if excel_data.get("strings"):
            base["strings"] = excel_data["strings"]
        if excel_data.get("monthly_consumption"):
            base["monthly_consumption"] = excel_data["monthly_consumption"]
        if excel_data.get("arrays"):
            base["arrays"] = excel_data["arrays"]
        if excel_data.get("dc_losses"):
            base["dc_losses"] = excel_data["dc_losses"]
        if excel_data.get("ac_losses"):
            base["ac_losses"] = excel_data["ac_losses"]
        if excel_data.get("balances_table"):
            base["balances_table"] = excel_data["balances_table"]
        if excel_data.get("loss_diagram"):
            base["loss_diagram"] = excel_data["loss_diagram"]
        if excel_data.get("p50_p90"):
            base["p50_p90"] = excel_data["p50_p90"]

    # Harmonisasi Spesifikasi PV Module (Astronergy vs JA Solar mengikut arahan drafter)
    pv_m_str = (str(base["pv_module"].get("model", "")) + " " + str(base["pv_module"].get("manufacturer", ""))).upper()
    watt_val = base["pv_module"].get("unit_power_w", 630)
    if "635" in pv_m_str or watt_val == 635:
        base["pv_module"]["manufacturer"] = "JA Solar"
        base["pv_module"]["model"] = "JAM72D42-635/LB"
        base["pv_module"]["pan_file"] = "JAM72D42-635/LB.PAN"
        base["pv_module"]["unit_power_w"] = 635
    elif "640" in pv_m_str or watt_val == 640:
        base["pv_module"]["manufacturer"] = "JA Solar"
        base["pv_module"]["model"] = "JAM72D42-640/LB"
        base["pv_module"]["pan_file"] = "JAM72D42-640/LB.PAN"
        base["pv_module"]["unit_power_w"] = 640
    else:
        base["pv_module"]["manufacturer"] = "Astronergy"
        base["pv_module"]["model"] = "ASTRO-N7 CHSM66RN-630W"
        base["pv_module"]["pan_file"] = "CHSM66RN(DG)F-BH-630.PAN"
        base["pv_module"]["unit_power_w"] = 630

    # Auto-fill andaian jika kosong
    client_key = base["client_name"] or base["short_address"] or "Project"
    if not base["glob_hor_kwh_m2"]:
        base["glob_hor_kwh_m2"] = get_year_input_for_state(base["state"], client_key)
    if not base["used_energy_kwh"]:
        base["used_energy_kwh"] = get_annual_load_for_model(base["inverter"]["model"], client_key, base["system_size_kwp"])
    if base["weather_data_variance"] is None:
        base["weather_data_variance"] = get_weather_assumption_for_state(base["state"])

    if base["strings"] and base["pv_module"]["total_panels"] == 0:
        base["pv_module"]["total_panels"] = sum(s.get("total_panels", 0) for s in base["strings"])

    return base

# ==============================================================================
# 5. PEMETAAN NILAI KE SEL KUNING (MAPPING ENGINE)
# ==============================================================================

def get_mapped_values_for_template(project_data):
    """Menghasilkan kamus pemetaan bagi setiap sel kuning: { ('SHEET', 'CELL'): mapped_value }"""
    p = project_data
    pv_m = p["pv_module"]
    inv = p["inverter"]
    bat = p["battery"]
    oris = p.get("orientations", [])
    strs = p.get("strings", [])

    ori_1 = oris[0]["str_val"] if len(oris) >= 1 else "20/180°"
    ori_2 = oris[1]["str_val"] if len(oris) >= 2 else "0/0"
    ori_3 = oris[2]["str_val"] if len(oris) >= 3 else "0/0"
    ori_4 = oris[3]["str_val"] if len(oris) >= 4 else "0/0"

    arr_1 = strs[0]["total_panels"] if len(strs) >= 1 else (pv_m.get("total_panels", 0) if ori_2 == "0/0" else 0)
    arr_2 = strs[1]["total_panels"] if len(strs) >= 2 else 0
    arr_3 = strs[2]["total_panels"] if len(strs) >= 3 else 0
    arr_4 = strs[3]["total_panels"] if len(strs) >= 4 else 0

    inv_kw = inv.get("unit_power_kw", 0.0)
    if isinstance(inv_kw, float) and inv_kw.is_integer():
        inv_kw = int(inv_kw)

    # Inverter model clean (NO HOYMILES)
    clean_inv_model = re.sub(r'(?i)\bHOYMILES\b|\bHUAWEI\b|\bSUNGROW\b', '', inv.get("model", "")).strip()

    mapped = {
        # Sheet PAGE 1
        ("PAGE 1", "B1"): p.get("client_name", ""),
        ("PAGE 1", "B2"): p.get("latitude", 5.233797),
        ("PAGE 1", "C2"): p.get("longitude", 100.448585),
        ("PAGE 1", "B3"): p.get("short_address", ""),
        ("PAGE 1", "B5"): int(p.get("glob_hor_kwh_m2", 1801)),
        ("PAGE 1", "B6"): int(p.get("used_energy_kwh", 16455)),

        # Sheet PAGE2
        ("PAGE2", "B9"): ori_1,
        ("PAGE2", "D9"): ori_2,
        ("PAGE2", "F9"): ori_3,
        ("PAGE2", "I9"): ori_4,
        ("PAGE2", "B13"): int(pv_m.get("total_panels", 0)),
        ("PAGE2", "D14"): inv_kw,
        ("PAGE2", "F13"): int(bat.get("units", 0) if bat.get("units") is not None else 0),

        # Sheet PAGE 3
        ("PAGE 3", "B14"): pv_m.get("manufacturer", "Astronergy"),
        ("PAGE 3", "B15"): pv_m.get("model", "ASTRO-N7 CHSM66RN-630W"),
        ("PAGE 3", "D15"): clean_inv_model,
        ("PAGE 3", "B16"): pv_m.get("pan_file", "CHSM66RN(DG)F-BH-630.PAN"),
        ("PAGE 3", "B17"): int(pv_m.get("unit_power_w", 630)),
        ("PAGE 3", "B22"): int(arr_1),
        ("PAGE 3", "B32"): int(arr_2),
        ("PAGE 3", "B42"): int(arr_3),
        ("PAGE 3", "B52"): int(arr_4),

        # Sheet PAGE 4 (DC Cable Lengths)
        ("PAGE 4", "B10"): 100 if arr_1 > 0 else 0,
        ("PAGE 4", "D10"): 80 if arr_2 > 0 else 0,
        ("PAGE 4", "B14"): 50 if arr_3 > 0 else 0,
        ("PAGE 4", "D14"): 80 if arr_4 > 0 else 0,

        # Sheet PAGE 7
        ("PAGE 7", "B2"): float(p.get("weather_data_variance", 3.8))
    }

    return mapped

# ==============================================================================
# 6. PENGISIAN SEL KUNING DALAM EXCEL
# ==============================================================================

def fill_yellow_cells(wb, project_data, custom_overrides=None):
    """Memasukkan data ke dalam sel-sel KUNING pada sheet PAGE 1 hingga PAGE 7."""
    mapped_vals = get_mapped_values_for_template(project_data)
    if custom_overrides and isinstance(custom_overrides, dict):
        mapped_vals.update(custom_overrides)

    for (s_name, coord), val in mapped_vals.items():
        if s_name in wb.sheetnames:
            ws = wb[s_name]
            try:
                cell = ws[coord]
                if cell.data_type == 'f':
                    continue
                ws[coord] = val
            except Exception:
                pass

# ==============================================================================
# 7. PEMILIHAN & PENJANAAN LAPORAN PVSYST PDF (6 TEMPLATE MATRIX)
# ==============================================================================

def select_pvsyst_template(project_data, template_dir=None):
    """
    Memilih fail template PVsyst PDF secara automatik mengikut matriks 6 template:
    1. Ada bateri (>0 unit: 1, 2, 3, 4, 5...) + 2 Orientasi -> BATTERY (2 Orientation).pdf
    2. Ada bateri (>0 unit: 1, 2, 3, 4, 5...) + 3 Orientasi -> BATTERY (3 Orientation).pdf
    3. Ada bateri (>0 unit: 1, 2, 3, 4, 5...) + 4 Orientasi -> BATTERY (4 Orientation).pdf
    4. Tiada bateri (0 / kosong / blank) + 2 Orientasi -> NO BATTERY (2 Orientation).pdf
    5. Tiada bateri (0 / kosong / blank) + 3 Orientasi -> NO BATTERY (3 Orientation).pdf
    6. Tiada bateri (0 / kosong / blank) + 4 Orientasi -> NO BATTERY (4 Orientation).pdf
    """
    if not template_dir:
        template_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in locals() and __file__ else os.getcwd()
    current_dir = template_dir
    
    # 1. Semak kewujudan bateri (>= 1 -> BATTERY, 0 / blank / None -> NO BATTERY)
    bat_count = project_data.get("battery_units")
    if bat_count is None and project_data.get("battery"):
        bat_count = project_data["battery"].get("units", 0)
    try:
        has_battery = bat_count is not None and float(bat_count) > 0
    except Exception:
        has_battery = False
    
    # 2. Semak bilangan orientasi
    orientations = project_data.get("orientations", [])
    active_orientations = [
        o for o in orientations 
        if str(o.get("str_val") or o.get("tilt_azimuth") or "").strip() not in ["0/0", "0/0°", "0", "None", "", "N/A", "0 / 0", "0 / 0°", "-", "0.0"]
    ]
    num_orientations = len(active_orientations)
    
    # Fallback to arrays if orientations not populated
    arrays = project_data.get("arrays", [])
    active_arrays = [a for a in arrays if (a.get("modules") or a.get("total_panels") or 0) > 0]
    if num_orientations == 0 and active_arrays:
        num_orientations = len(active_arrays)
        
    if num_orientations <= 2:
        ori_str = "2 Orientation"
    elif num_orientations == 3:
        ori_str = "3 Orientation"
    else:
        ori_str = "4 Orientation"
        
    prefix = "BATTERY" if has_battery else "NO BATTERY"
    selected = f"{prefix} ({ori_str}).pdf"

    selected_path = os.path.join(current_dir, selected)
    if os.path.exists(selected_path):
        return selected_path
    elif os.path.exists(selected):
        return selected
        # Fallback
        return os.path.join(current_dir, f"{prefix} (2 Orientation).pdf")


def generate_pvsyst_pdf(template_pdf_source, excel_data):
    """
    Menjana Laporan PVsyst PDF yang 100% tepat mengikut template asal tanpa mengubah struktur/reka letak template:
    1. Menggantikan nombor dan teks mengikut fail Excel secara tepat.
    2. Sifar pertindihan teks (Zero overlap) dengan background mask putih yang bersih.
    3. Logo PVsyst asal terpelihara 100%.
    4. Header Project & Variant kemas dan berpusat (#0d307d).
    5. Menghapuskan 100% semua highlight anotasi kuning.
    """
    current_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in locals() and __file__ else os.getcwd()
    
    if not template_pdf_source or template_pdf_source == "auto":
        template_pdf_source = select_pvsyst_template(excel_data)

    if isinstance(template_pdf_source, (bytes, bytearray)):
        doc = pymupdf.open(stream=template_pdf_source, filetype="pdf")
    elif isinstance(template_pdf_source, io.BytesIO):
        doc = pymupdf.open(stream=template_pdf_source.getvalue(), filetype="pdf")
    else:
        if not os.path.exists(template_pdf_source):
            cand_path = os.path.join(current_dir, template_pdf_source)
            if os.path.exists(cand_path):
                template_pdf_source = cand_path
            else:
                template_pdf_source = select_pvsyst_template(excel_data)
        doc = pymupdf.open(template_pdf_source)

    # 1. Pembolehubah Excel
    client_name = str(excel_data.get("client_name") or "").strip()
    short_address = str(excel_data.get("short_address") or "").strip()
    lat_val = f"{float(excel_data.get('latitude') or 5.2338):.4f}"
    lon_val = f"{float(excel_data.get('longitude') or 100.4486):.4f}"
    
    kwp_val = float(excel_data.get("system_size_kwp") or excel_data.get("kwp") or 12.60)
    variant_str = f"{kwp_val:.2f}"
    
    kwac_num = float(excel_data.get("system_size_kwac") or excel_data.get("kwac") or 12.0)
    kwac_str = f"{int(kwac_num)}" if kwac_num.is_integer() else f"{kwac_num:.1f}"
    
    total_mods = int(excel_data.get("total_modules") or (excel_data.get("pv_module", {}).get("total_panels") if excel_data.get("pv_module") else 20) or 20)
    inv_units = int(excel_data.get("inverter_units") or (excel_data.get("inverter", {}).get("total_units") if excel_data.get("inverter") else 1) or 1)
    
    pnom_rat = float(excel_data.get("pnom_ratio") or (excel_data.get("inverter", {}).get("pnom_ratio") if excel_data.get("inverter") and excel_data.get("inverter", {}).get("pnom_ratio") else (kwp_val / kwac_num if kwac_num > 0 else 1.26)))
    
    orientations = excel_data.get("orientations", [])
    def get_ori_str(idx, default="0/0"):
        if idx < len(orientations):
            o = orientations[idx]
            val = str(o.get("str_val") or o.get("tilt_azimuth") or default).strip()
            return val if val else default
        return str(excel_data.get(f"ori{idx+1}", default)).strip()

    ori1 = get_ori_str(0, "20/103°")
    ori2 = get_ori_str(1, "20/-167°")
    ori3 = get_ori_str(2, "0/0")
    ori4 = get_ori_str(3, "0/0")
    
    def parse_tilt_az(s):
        s_clean = s.replace("°", "").replace("deg", "").strip()
        parts = s_clean.split("/")
        if len(parts) >= 2:
            return parts[0].strip(), parts[1].strip()
        return "20", "180"

    t1, a1 = parse_tilt_az(ori1)
    t2, a2 = parse_tilt_az(ori2)
    t3, a3 = parse_tilt_az(ori3)
    t4, a4 = parse_tilt_az(ori4)

    pv_mfg = str(excel_data.get("pv_module_mfg") or (excel_data.get("pv_module", {}).get("manufacturer") if excel_data.get("pv_module") else "Astronergy")).strip()
    pv_model = str(excel_data.get("pv_module_model") or (excel_data.get("pv_module", {}).get("model") if excel_data.get("pv_module") else "ASTRO-N7 CHSM66RN-630W")).strip()
    pv_pan = str(excel_data.get("pv_module_pan") or (excel_data.get("pv_module", {}).get("pan_file") if excel_data.get("pv_module") else "CHSM66RN(DG)F-BH-630.PAN")).strip()
    pv_pwr = int(excel_data.get("pv_module_power") or (excel_data.get("pv_module", {}).get("unit_power_w") if excel_data.get("pv_module") else 630))

    inv_mfg = str(excel_data.get("inverter_mfg") or (excel_data.get("inverter", {}).get("manufacturer") if excel_data.get("inverter") else "Hoymiles")).strip()
    raw_inv = str(excel_data.get("inverter_model") or (excel_data.get("inverter", {}).get("model") if excel_data.get("inverter") else "HIS-5.0L-G3")).strip()
    inv_model = re.sub(r'(?i)\bHOYMILES\b', '', raw_inv).strip()
    inv_pwr = float(excel_data.get("inverter_power") or (excel_data.get("inverter", {}).get("unit_power_kw") if excel_data.get("inverter") else kwac_num) or kwac_num)

    bat_units = int(excel_data.get("battery_units") or (excel_data.get("battery", {}).get("units") if excel_data.get("battery") and excel_data.get("battery", {}).get("units") is not None else 0))
    has_bat = bat_units > 0
    
    bat_volt = 51
    try:
        raw_bv = excel_data.get("battery_voltage") or (excel_data.get("battery", {}).get("voltage_v") or excel_data.get("battery", {}).get("voltage") if excel_data.get("battery") else None)
        if raw_bv is not None:
            bat_volt = int(float(raw_bv))
    except Exception:
        bat_volt = 51

    bat_cap = 314
    try:
        raw_bc = excel_data.get("battery_capacity") or (excel_data.get("battery", {}).get("capacity_ah") or excel_data.get("battery", {}).get("capacity") if excel_data.get("battery") else None)
        if raw_bc is not None:
            bat_cap = int(float(raw_bc))
    except Exception:
        bat_cap = 314

    bat_model = str(excel_data.get("battery_model") or (excel_data.get("battery", {}).get("model") if excel_data.get("battery") and excel_data.get("battery", {}).get("model") else "LB-16D-G3")).strip()
    has_bat = bat_units > 0

def generate_pvsyst_pdf(template_pdf_source, excel_data):
    """
    Menjana Laporan PVsyst PDF yang 100% tepat mengikut template asal tanpa mengubah struktur/reka letak template:
    1. Menggantikan nombor dan teks mengikut fail Excel secara tepat.
    2. Sifar pertindihan teks (Zero overlap) dengan background mask putih yang bersih.
    3. Logo PVsyst asal terpelihara 100%.
    4. Header Project & Variant kemas dan berpusat (#0d307d).
    5. Menghapuskan 100% semua highlight anotasi kuning.
    """
    current_dir = os.path.dirname(os.path.abspath(__file__)) if "__file__" in locals() and __file__ else os.getcwd()
    
    if not template_pdf_source or template_pdf_source == "auto":
        template_pdf_source = select_pvsyst_template(excel_data)

    if isinstance(template_pdf_source, (bytes, bytearray)):
        doc = pymupdf.open(stream=template_pdf_source, filetype="pdf")
    elif isinstance(template_pdf_source, io.BytesIO):
        doc = pymupdf.open(stream=template_pdf_source.getvalue(), filetype="pdf")
    else:
        if not os.path.exists(template_pdf_source):
            cand_path = os.path.join(current_dir, template_pdf_source)
            if os.path.exists(cand_path):
                template_pdf_source = cand_path
            else:
                template_pdf_source = select_pvsyst_template(excel_data)
        doc = pymupdf.open(template_pdf_source)

    # 1. Pembolehubah Excel
    client_name = str(excel_data.get("client_name") or "").strip()
    short_address = str(excel_data.get("short_address") or "").strip()
    lat_val = f"{float(excel_data.get('latitude') or 5.2338):.4f}"
    lon_val = f"{float(excel_data.get('longitude') or 100.4486):.4f}"
    
    kwp_val = float(excel_data.get("system_size_kwp") or excel_data.get("kwp") or 12.60)
    variant_str = f"{kwp_val:.2f}"
    
    kwac_num = float(excel_data.get("system_size_kwac") or excel_data.get("kwac") or 12.0)
    kwac_str = f"{int(kwac_num)}" if kwac_num.is_integer() else f"{kwac_num:.1f}"
    
    total_mods = int(excel_data.get("total_modules") or (excel_data.get("pv_module", {}).get("total_panels") if excel_data.get("pv_module") else 20) or 20)
    inv_units = int(excel_data.get("inverter_units") or (excel_data.get("inverter", {}).get("total_units") if excel_data.get("inverter") else 1) or 1)
    
    pnom_rat = float(excel_data.get("pnom_ratio") or (excel_data.get("inverter", {}).get("pnom_ratio") if excel_data.get("inverter") and excel_data.get("inverter", {}).get("pnom_ratio") else (kwp_val / kwac_num if kwac_num > 0 else 1.26)))
    
    orientations = excel_data.get("orientations", [])
    def get_ori_str(idx, default="0/0"):
        if idx < len(orientations):
            o = orientations[idx]
            val = str(o.get("str_val") or o.get("tilt_azimuth") or default).strip()
            return val if val else default
        return str(excel_data.get(f"ori{idx+1}", default)).strip()

    ori1 = get_ori_str(0, "20/103°")
    ori2 = get_ori_str(1, "20/-167°")
    ori3 = get_ori_str(2, "0/0")
    ori4 = get_ori_str(3, "0/0")
    
    def parse_tilt_az(s):
        s_clean = s.replace("°", "").replace("deg", "").strip()
        parts = s_clean.split("/")
        if len(parts) >= 2:
            return parts[0].strip(), parts[1].strip()
        return "20", "180"

    t1, a1 = parse_tilt_az(ori1)
    t2, a2 = parse_tilt_az(ori2)
    t3, a3 = parse_tilt_az(ori3)
    t4, a4 = parse_tilt_az(ori4)

    pv_mfg = str(excel_data.get("pv_module_mfg") or (excel_data.get("pv_module", {}).get("manufacturer") if excel_data.get("pv_module") else "Astronergy")).strip()
    pv_model = str(excel_data.get("pv_module_model") or (excel_data.get("pv_module", {}).get("model") if excel_data.get("pv_module") else "ASTRO-N7 CHSM66RN-630W")).strip()
    pv_pan = str(excel_data.get("pv_module_pan") or (excel_data.get("pv_module", {}).get("pan_file") if excel_data.get("pv_module") else "CHSM66RN(DG)F-BH-630.PAN")).strip()
    pv_pwr = int(excel_data.get("pv_module_power") or (excel_data.get("pv_module", {}).get("unit_power_w") if excel_data.get("pv_module") else 630))

    inv_mfg = str(excel_data.get("inverter_mfg") or (excel_data.get("inverter", {}).get("manufacturer") if excel_data.get("inverter") else "Hoymiles")).strip()
    raw_inv = str(excel_data.get("inverter_model") or (excel_data.get("inverter", {}).get("model") if excel_data.get("inverter") else "HIS-5.0L-G3")).strip()
    inv_model = re.sub(r'(?i)\bHOYMILES\b', '', raw_inv).strip()
    inv_pwr = float(excel_data.get("inverter_power") or (excel_data.get("inverter", {}).get("unit_power_kw") if excel_data.get("inverter") else kwac_num) or kwac_num)

    bat_units = int(excel_data.get("battery_units") or (excel_data.get("battery", {}).get("units") if excel_data.get("battery") and excel_data.get("battery", {}).get("units") is not None else 0))
    has_bat = bat_units > 0
    
    bat_volt = 51
    try:
        raw_bv = excel_data.get("battery_voltage") or (excel_data.get("battery", {}).get("voltage_v") or excel_data.get("battery", {}).get("voltage") if excel_data.get("battery") else None)
        if raw_bv is not None:
            bat_volt = int(float(raw_bv))
    except Exception:
        bat_volt = 51

    bat_cap = 314
    try:
        raw_bc = excel_data.get("battery_capacity") or (excel_data.get("battery", {}).get("capacity_ah") or excel_data.get("battery", {}).get("capacity") if excel_data.get("battery") else None)
        if raw_bc is not None:
            bat_cap = int(float(raw_bc))
    except Exception:
        bat_cap = 314

    bat_model = str(excel_data.get("battery_model") or (excel_data.get("battery", {}).get("model") if excel_data.get("battery") and excel_data.get("battery", {}).get("model") else "LB-16D-G3")).strip()
    has_bat = bat_units > 0

    prod_kwh = int(round(float(excel_data.get("produced_energy_kwh") or 14975)))
    spec_prod_raw = float(excel_data.get("specific_prod_kwh_kwp") or excel_data.get("specific_production") or (prod_kwh / kwp_val if kwp_val > 0 else 1188.5))
    spec_str = f"{spec_prod_raw:.1f}" if not spec_prod_raw.is_integer() else f"{int(spec_prod_raw)}"
    
    pr_pct = float(excel_data.get("perf_ratio_pr") or 81.45)
    used_kwh = int(round(float(excel_data.get("used_energy_kwh") or 10250)))
    sf_pct = float(excel_data.get("solar_fraction_sf") or 100.0)

    m_vals = excel_data.get("monthly_consumption", [])
    arrays_data = excel_data.get("arrays", [])
    
    dc_l = excel_data.get("dc_losses", {"arr1_res": 292, "arr1_loss": 1.32, "arr2_res": 233, "arr2_loss": 1.90, "arr3_res": 146, "arr3_loss": 0.74, "arr4_res": 0, "arr4_loss": 0.0})
    is_hit_model = "HIT" in raw_inv.upper() or kwac_num >= 8.0
    ac_l = excel_data.get("ac_losses", {
        "inverter_voltage": f"{400 if is_hit_model else 220} Vac" + (" tri" if is_hit_model else " mono"),
        "wire_section": f"Copper {'3 x 3' if is_hit_model else '2 x 4'} mm²",
        "wires_length": f"{20 if is_hit_model else 10} m"
    })
    
    # 8. Loss Diagram values from PAGE 6
    loss_diag = excel_data.get("loss_diagram", {})
    gh_v = float(loss_diag.get("glob_hor") or excel_data.get("glob_hor_kwh_m2") or 1608.9)
    gi_v = float(loss_diag.get("glob_inc") or 1549.0)
    pvn_v = float(loss_diag.get("pv_nom") or 10572.6)
    av_v = float(loss_diag.get("array_virt") or 9674.0)
    fe_v = float(loss_diag.get("final_energy") or 9362.0)
    ef_v = float(loss_diag.get("efr_grid") or 0.0)
    bs_v = float(loss_diag.get("battery_stored") or 6400.0)
    du_v = float(loss_diag.get("direct_use") or 2572.62)

    # 9. P50-P90 values from PAGE 7
    p50_info = excel_data.get("p50_p90", {})
    var_v = float(excel_data.get("weather_data_variance") or p50_info.get("variance") or 3.8)
    cvar_v = float(p50_info.get("custom_variability") or var_v)
    p_var_n = int(excel_data.get("p50_variability_kwh") or p50_info.get("variability") or 666)
    p50_n = int(excel_data.get("p50_kwh") or p50_info.get("p50") or 10573)
    p90_n = int(excel_data.get("p90_kwh") or p50_info.get("p90") or 10159)
    p75_n = int(excel_data.get("p75_kwh") or p50_info.get("p75") or 9786)
    egrid_n = p50_n

    pw = 595.3
    HEADER_BLUE = (13/255, 48/255, 125/255)

    for p_idx, page in enumerate(doc):
        p_txt = page.get_text("text")
        
        redacts = []
        inserts = []

        def add_redact(rect):
            if rect:
                redacts.append(rect)

        def add_text(pos, text, fontsize=7.5, fontname="helv", color=(0, 0, 0)):
            inserts.append((pos, text, fontsize, fontname, color))

        # 1. Header (Pages 2+)
        if p_idx > 0:
            add_redact(pymupdf.Rect(80, 24, 520, 46))
            th1 = f"Project: {client_name}"
            wh1 = pymupdf.get_text_length(th1, fontname="helv", fontsize=12.0)
            add_text(((pw - wh1) / 2, 38), th1, fontsize=12.0, fontname="helv", color=HEADER_BLUE)

            add_redact(pymupdf.Rect(150, 48, 450, 66))
            th2 = f"Variant: {variant_str}"
            wh2 = pymupdf.get_text_length(th2, fontname="helv", fontsize=10.0)
            add_text(((pw - wh2) / 2, 58), th2, fontsize=10.0, fontname="helv", color=HEADER_BLUE)

        # 2. Cover Page (Page 1)
        if p_idx == 0:
            add_redact(pymupdf.Rect(40, 210, 555, 235))
            t1_cov = f"Project: {client_name}"
            w1_cov = pymupdf.get_text_length(t1_cov, fontname="helv", fontsize=14.0)
            add_text(((pw - w1_cov) / 2, 227), t1_cov, fontsize=14.0, fontname="helv", color=(0, 0, 0))

            add_redact(pymupdf.Rect(40, 236, 555, 255))
            t2_cov = f"Variant: {variant_str}"
            w2_cov = pymupdf.get_text_length(t2_cov, fontname="helv", fontsize=12.0)
            add_text(((pw - w2_cov) / 2, 248), t2_cov, fontsize=12.0, fontname="helv", color=(0, 0, 0))

            add_redact(pymupdf.Rect(40, 271, 555, 289))
            t3_cov = f"System power: {variant_str} kWp"
            w3_cov = pymupdf.get_text_length(t3_cov, fontname="helv", fontsize=12.0)
            add_text(((pw - w3_cov) / 2, 284), t3_cov, fontsize=12.0, fontname="helv", color=(0, 0, 0))

            add_redact(pymupdf.Rect(40, 290, 555, 310))
            t4_cov = f"{short_address} - Malaysia"
            w4_cov = pymupdf.get_text_length(t4_cov, fontname="helv", fontsize=12.0)
            add_text(((pw - w4_cov) / 2, 302), t4_cov, fontsize=12.0, fontname="helv", color=(0, 0, 0))

        # 3. Monthly values table on Page 3
        jan_m = [r for r in page.search_for("Jan.") if 150 < r.y0 < 360] or [r for r in page.search_for("Jan") if 150 < r.y0 < 360]
        if jan_m and len(m_vals) >= 13:
            jr = jan_m[0]
            add_redact(pymupdf.Rect(43.0, jr.y1 + 1.0, 505.0, jr.y1 + 13.5))
            col_centers = [57.5, 93.2, 128.8, 164.5, 200.1, 235.8, 271.4, 307.0, 342.7, 378.3, 414.0, 449.6, 485.2]
            for idx_m in range(13):
                v_m = m_vals[idx_m]
                v_str_m = str(int(round(float(v_m))))
                w_m = pymupdf.get_text_length(v_str_m, fontname="helv", fontsize=6.8)
                cx_m = col_centers[idx_m] - (w_m / 2.0)
                add_text((cx_m, jr.y1 + 10.0), v_str_m, fontsize=6.8, fontname="helv", color=(0, 0, 0))

        # 4. Page 2 System Information & Results Summary
        if p_idx == 1:
            site_m = page.search_for("Geographical Site")
            if site_m:
                sr = site_m[0]
                add_redact(pymupdf.Rect(40, sr.y1 + 1, 280, sr.y1 + 12))
                add_text((42.5, sr.y1 + 9), short_address, fontsize=7.5, fontname="helv", color=(0, 0, 0))

            w_m = page.search_for("Weather data")
            if w_m:
                wr = w_m[0]
                add_redact(pymupdf.Rect(40, wr.y1 + 1, 280, wr.y1 + 12))
                add_text((42.5, wr.y1 + 9), short_address, fontsize=7.5, fontname="helv", color=(0, 0, 0))

            sit_m = page.search_for("Situation")
            if sit_m:
                sitr = sit_m[0]
                add_redact(pymupdf.Rect(300, sitr.y1 + 1, 380, sitr.y1 + 25))
                add_text((305, sitr.y1 + 10), f"{lat_val} °(N)", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((305, sitr.y1 + 22), f"{lon_val} °(E)", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            # Orientations Page 2
            add_redact(pymupdf.Rect(115, 320, 175, 336))
            add_text((125, 332.5), f"{t1} / {a1} °", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            add_redact(pymupdf.Rect(295, 320, 355, 336))
            add_text((305, 332.5), f"{t2} / {a2} °", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            if "Orientation #3" in p_txt and ori3 not in ["0/0", "0/0°", "0", ""]:
                add_redact(pymupdf.Rect(475, 320, 535, 336))
                add_text((485, 332.5), f"{t3} / {a3} °", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            if "Orientation #4" in p_txt and ori4 not in ["0/0", "0/0°", "0", ""]:
                add_redact(pymupdf.Rect(115, 350, 175, 365))
                add_text((125, 361.5), f"{t4} / {a4} °", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            # PV Array Page 2
            mods_m = [r for r in page.search_for("Nb. of modules") if r.x0 < 150]
            if mods_m:
                mr = mods_m[0]
                if has_bat:
                    add_redact(pymupdf.Rect(120, mr.y0 - 2, 161.0, mr.y1 + 2))
                    w_tm = pymupdf.get_text_length(str(total_mods), fontname="helv", fontsize=7.5)
                    add_text((159.0 - w_tm, mr.y1 - 1), str(total_mods), fontsize=7.5, fontname="helv", color=(0, 0, 0))
                else:
                    add_redact(pymupdf.Rect(180, mr.y0 - 2, 227.0, mr.y1 + 2))
                    w_tm = pymupdf.get_text_length(str(total_mods), fontname="helv", fontsize=7.5)
                    add_text((225.5 - w_tm, mr.y1 - 1), str(total_mods), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            pnom_m = [r for r in page.search_for("Pnom total") if r.x0 < 150]
            if pnom_m:
                pr = pnom_m[0]
                if has_bat:
                    add_redact(pymupdf.Rect(120, pr.y0 - 2, 161.0, pr.y1 + 2))
                    w_pn = pymupdf.get_text_length(variant_str, fontname="helv", fontsize=7.5)
                    add_text((159.0 - w_pn, pr.y1 - 1), variant_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                else:
                    add_redact(pymupdf.Rect(180, pr.y0 - 2, 227.0, pr.y1 + 2))
                    w_pn = pymupdf.get_text_length(variant_str, fontname="helv", fontsize=7.5)
                    add_text((225.5 - w_pn, pr.y1 - 1), variant_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))

            # Inverters Page 2
            inv_u_m = [r for r in page.search_for("Nb. of units") if (200 < r.x0 < 350 if has_bat else 300 < r.x0 < 400)]
            if inv_u_m:
                ir = inv_u_m[0]
                if has_bat:
                    add_redact(pymupdf.Rect(300, ir.y0 - 2, 340.0, ir.y1 + 2))
                    w_iu = pymupdf.get_text_length(str(inv_units), fontname="helv", fontsize=7.5)
                    add_text((339.0 - w_iu, ir.y1 - 1), str(inv_units), fontsize=7.5, fontname="helv", color=(0, 0, 0))
                else:
                    add_redact(pymupdf.Rect(460, ir.y0 - 2, 496.0, ir.y1 + 2))
                    w_iu = pymupdf.get_text_length(str(inv_units), fontname="helv", fontsize=7.5)
                    add_text((494.5 - w_iu, ir.y1 - 1), str(inv_units), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            tot_p_m = [r for r in page.search_for("Total power") if (200 < r.x0 < 350 if has_bat else 300 < r.x0 < 400)]
            if tot_p_m:
                tpr = tot_p_m[0]
                if has_bat:
                    add_redact(pymupdf.Rect(300, tpr.y0 - 2, 340.0, tpr.y1 + 2))
                    w_tp = pymupdf.get_text_length(kwac_str, fontname="helv", fontsize=7.5)
                    add_text((339.0 - w_tp, tpr.y1 - 1), kwac_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                else:
                    add_redact(pymupdf.Rect(460, tpr.y0 - 2, 496.0, tpr.y1 + 2))
                    w_tp = pymupdf.get_text_length(kwac_str, fontname="helv", fontsize=7.5)
                    add_text((494.5 - w_tp, tpr.y1 - 1), kwac_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))

            pnom_r_m = [r for r in page.search_for("Pnom ratio") if (200 < r.x0 < 350 if has_bat else 300 < r.x0 < 400)]
            if pnom_r_m:
                prr = pnom_r_m[0]
                if has_bat:
                    add_redact(pymupdf.Rect(300, prr.y0 - 2, 340.0, prr.y1 + 2))
                    w_prr = pymupdf.get_text_length(f"{pnom_rat:.2f}", fontname="helv", fontsize=7.5)
                    add_text((339.0 - w_prr, prr.y1 - 1), f"{pnom_rat:.2f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                else:
                    add_redact(pymupdf.Rect(460, prr.y0 - 2, 496.0, prr.y1 + 2))
                    w_prr = pymupdf.get_text_length(f"{pnom_rat:.2f}", fontname="helv", fontsize=7.5)
                    add_text((494.5 - w_prr, prr.y1 - 1), f"{pnom_rat:.2f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            # Battery pack Page 2
            if has_bat:
                bat_u_m = [r for r in page.search_for("Nb. of units") if r.x0 > 380]
                if bat_u_m:
                    bur = bat_u_m[0]
                    add_redact(pymupdf.Rect(495, bur.y0 - 2, 520, bur.y1 + 2))
                    w_bu = pymupdf.get_text_length(str(bat_units), fontname="helv", fontsize=7.5)
                    add_text((518 - w_bu, bur.y1 - 1), str(bat_units), fontsize=7.5, fontname="helv", color=(0, 0, 0))

                volt_m = [r for r in page.search_for("Voltage") if r.x0 > 380]
                if volt_m:
                    vr = volt_m[0]
                    add_redact(pymupdf.Rect(495, vr.y0 - 2, 520, vr.y1 + 2))
                    w_bv = pymupdf.get_text_length(str(bat_volt), fontname="helv", fontsize=7.5)
                    add_text((518 - w_bv, vr.y1 - 1), str(bat_volt), fontsize=7.5, fontname="helv", color=(0, 0, 0))

                cap_m = [r for r in page.search_for("Capacity") if r.x0 > 380]
                if cap_m:
                    cr = cap_m[0]
                    add_redact(pymupdf.Rect(495, cr.y0 - 2, 520, cr.y1 + 2))
                    w_bc = pymupdf.get_text_length(str(bat_cap), fontname="helv", fontsize=7.5)
                    add_text((518 - w_bc, cr.y1 - 1), str(bat_cap), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            # Results summary Page 2
            pe_m = page.search_for("Produced Energy")
            if pe_m:
                per = pe_m[0]
                add_redact(pymupdf.Rect(110, per.y0 - 2, 161.0, per.y1 + 2))
                w_pe = pymupdf.get_text_length(str(prod_kwh), fontname="helv", fontsize=7.5)
                add_text((160 - w_pe, per.y1 - 1), str(prod_kwh), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            ue_m = page.search_for("Used Energy")
            if ue_m and has_bat:
                uer = ue_m[0]
                add_redact(pymupdf.Rect(110, uer.y0 - 2, 161.0, uer.y1 + 2))
                w_ue = pymupdf.get_text_length(str(used_kwh), fontname="helv", fontsize=7.5)
                add_text((160 - w_ue, uer.y1 - 1), str(used_kwh), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            sp_m = page.search_for("Specific prod.") or page.search_for("Specific production")
            if sp_m:
                spr = sp_m[0]
                add_redact(pymupdf.Rect(295, spr.y0 - 2, 340.0, spr.y1 + 2))
                w_sp = pymupdf.get_text_length(spec_str, fontname="helv", fontsize=7.5)
                add_text((339 - w_sp, spr.y1 - 1), spec_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))

            sf_m = page.search_for("Solar Fraction")
            if sf_m and has_bat:
                sfr = sf_m[0]
                add_redact(pymupdf.Rect(465, sfr.y0 - 2, 520.0, sfr.y1 + 2))
                w_sf = pymupdf.get_text_length(f"{sf_pct:.1f}", fontname="helv", fontsize=7.5)
                add_text((519 - w_sf, sfr.y1 - 1), f"{sf_pct:.1f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            pr_m = page.search_for("Perf. Ratio")
            if pr_m:
                pr_r = pr_m[0]
                add_redact(pymupdf.Rect(465, pr_r.y0 - 2, 520.0, pr_r.y1 + 2))
                w_pr = pymupdf.get_text_length(f"{pr_pct:.2f}", fontname="helv", fontsize=7.5)
                add_text((519 - w_pr, pr_r.y1 - 1), f"{pr_pct:.2f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

        # 5. General parameters on Page 3 (p_idx >= 2)
        if p_idx >= 2 and "General parameters" in p_txt:
            add_redact(pymupdf.Rect(125, 208, 175, 222))
            add_text((135, 219.0), f"{t1} / {a1} °", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            add_redact(pymupdf.Rect(300, 208, 355, 222))
            add_text((310, 219.0), f"{t2} / {a2} °", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            if "Orientation #3" in p_txt and ori3 not in ["0/0", "0/0°", "0", ""]:
                add_redact(pymupdf.Rect(475, 208, 535, 222))
                add_text((485, 219.0), f"{t3} / {a3} °", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            if "Orientation #4" in p_txt and ori4 not in ["0/0", "0/0°", "0", ""]:
                add_redact(pymupdf.Rect(125, 250, 180, 265))
                add_text((130, 261.7), f"{t4} / {a4} °", fontsize=7.5, fontname="helv", color=(0, 0, 0))

        # 6. PV Module & Inverter Blocks (ONLY on Page 3 with General parameters)
        if p_idx >= 2 and "General parameters" in p_txt and "PV Array Characteristics" in p_txt:
            pvm_m = [r for r in page.search_for("PV module") if r.x0 < 60]
            if pvm_m:
                r_pv = pvm_m[0]
                add_redact(pymupdf.Rect(45, r_pv.y0 + 10, 260, r_pv.y0 + 72))
                add_text((42.5, r_pv.y0 + 20), "Manufacturer", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((195, r_pv.y0 + 20), pv_mfg, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((42.5, r_pv.y0 + 32), "Model", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((115, r_pv.y0 + 32), pv_model, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((55, r_pv.y0 + 44), "(Custom parameters definition)", fontsize=7.0, fontname="helv", color=(0, 0, 0))
                add_text((55, r_pv.y0 + 56), pv_pan, fontsize=7.0, fontname="helv", color=(0, 0, 0))
                add_text((42.5, r_pv.y0 + 68), "Unit Nom. Power", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((205, r_pv.y0 + 68), f"{pv_pwr} Wp", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            inv_m = [r for r in page.search_for("Inverter") if 300 < r.x0 < 350 and r.y0 < 450]
            if inv_m:
                r_inv = inv_m[0]
                add_redact(pymupdf.Rect(310, r_inv.y0 + 10, 545, r_inv.y0 + 72))
                add_text((311.9, r_inv.y0 + 20), "Manufacturer", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((465, r_inv.y0 + 20), inv_mfg, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((311.9, r_inv.y0 + 32), "Model", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((440, r_inv.y0 + 32), inv_model, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((325, r_inv.y0 + 44), "(Custom parameters definition)", fontsize=7.0, fontname="helv", color=(0, 0, 0))
                add_text((311.9, r_inv.y0 + 68), "Unit Nom. Power", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((475, r_inv.y0 + 68), f"{inv_pwr:.1f} kWac", fontsize=7.5, fontname="helv", color=(0, 0, 0))

        # 7. Sub-arrays 1..4 (p_idx >= 2)
        if p_idx >= 2 and "PV Array Characteristics" in p_txt:
            def render_array_sec(arr_title, arr_data, ori_idx, def_tilt_az):
                arr_m = page.search_for(arr_title)
                if not arr_m:
                    return
                r_arr = None
                for candidate in arr_m:
                    clip_txt = page.get_text("text", clip=pymupdf.Rect(candidate.x0, candidate.y0, candidate.x0 + 150, candidate.y0 + 35))
                    if "Orientation" in clip_txt or "Tilt" in clip_txt or "modules" in clip_txt:
                        r_arr = candidate
                        break
                if not r_arr:
                    return
                if not arr_data or (arr_data.get("modules", 0) == 0 and arr_data.get("total_panels", 0) == 0):
                    return

                t_val, a_val = parse_tilt_az(get_ori_str(ori_idx, def_tilt_az))
                mods_cnt = arr_data.get("modules") or arr_data.get("total_panels") or 5
                stc_v = arr_data.get("stc_wp") or (mods_cnt * pv_pwr)
                pmpp_v = arr_data.get("pmpp") or round(stc_v * 0.933)
                umpp_v = arr_data.get("umpp") or round(mods_cnt * 37.7)
                impp_v = arr_data.get("impp") or 16
                
                raw_mppt = arr_data.get("mppt_pct")
                if raw_mppt is None or raw_mppt == "":
                    raw_mppt = round((mods_cnt / total_mods) * 100) if total_mods > 0 else 50
                if isinstance(raw_mppt, str):
                    mppt_clean = raw_mppt.replace("%", "").strip()
                    try:
                        mppt_num = float(mppt_clean)
                    except Exception:
                        mppt_num = 25.0
                else:
                    try:
                        mppt_num = float(raw_mppt)
                    except Exception:
                        mppt_num = 25.0
                if 0 < mppt_num <= 1.0:
                    mppt_num = mppt_num * 100.0
                mppt_pct = f"{int(round(mppt_num))}%"
                
                inv_frac_raw = arr_data.get("inv_unit_fraction")
                if inv_frac_raw is None or inv_frac_raw == "":
                    inv_frac_val = round(mods_cnt / total_mods, 2) if total_mods > 0 else 0.5
                    inv_frac = f"{inv_frac_val:.2f}".rstrip("0").rstrip(".") + " unit"
                elif isinstance(inv_frac_raw, (int, float)):
                    inv_frac = f"{float(inv_frac_raw):.2f}".rstrip("0").rstrip(".") + " unit"
                else:
                    inv_frac = str(inv_frac_raw)
                    if not inv_frac.endswith("unit"):
                        inv_frac += " unit"
                
                tot_p_raw = arr_data.get("tot_pwr") or arr_data.get("inverter_power_kwac")
                if tot_p_raw is None or tot_p_raw == "":
                    tot_p_num = kwac_num * (mods_cnt / total_mods) if total_mods > 0 else kwac_num
                    tot_p = f"{tot_p_num:.1f} kWac"
                else:
                    try:
                        tot_p_num = float(str(tot_p_raw).replace("kWac", "").strip())
                        tot_p = f"{tot_p_num:.1f} kWac"
                    except Exception:
                        tot_p = str(tot_p_raw)
                
                op_v = str(arr_data.get("operating_voltage") or ("150-900 V" if is_hit_model else "90-550 V"))
                if not op_v.endswith("V"):
                    op_v = f"{op_v} V"
                pnom_r_raw = arr_data.get("pnom_ratio")
                if pnom_r_raw is None or pnom_r_raw == "":
                    pnom_r_raw = pnom_rat
                try:
                    pnom_r = f"{float(pnom_r_raw):.2f}"
                except Exception:
                    pnom_r = str(pnom_r_raw)

                clip_blk = pymupdf.Rect(r_arr.x0, r_arr.y0, 550, r_arr.y0 + 130)
                def get_lbl_y(label_str, fallback_offset):
                    m = page.search_for(label_str, clip=clip_blk)
                    if m:
                        return m[0].y1 - 1
                    return r_arr.y0 + fallback_offset

                y_ori = get_lbl_y("Orientation", 16)
                y_ta = get_lbl_y("Tilt/Azimuth", 28)
                y_nm = get_lbl_y("Number of PV modules", 40)
                y_stc = get_lbl_y("Nominal (STC)", 52)
                y_mod = y_stc + 12.0  # Modules is directly below Nominal STC
                y_pmpp = get_lbl_y("Pmpp", 88)
                y_umpp = get_lbl_y("U mpp", 100)
                y_impp = get_lbl_y("I mpp", 112)

                y_ni = get_lbl_y("Number of inverters", 40)
                y_tp = get_lbl_y("Total power", 52)
                y_op = get_lbl_y("Operating voltage", 76)
                y_pr = get_lbl_y("Pnom ratio", 88)

                # Left PV side (clean redaction)
                add_redact(pymupdf.Rect(160, r_arr.y0 + 5, 260, r_arr.y0 + 125))
                w_ori = pymupdf.get_text_length(str(ori_idx + 1), fontname="helv", fontsize=7.5)
                add_text((225.5 - w_ori, y_ori), str(ori_idx + 1), fontsize=7.5, fontname="helv", color=(0, 0, 0))

                w_ta = pymupdf.get_text_length(f"{t_val}/{a_val}", fontname="helv", fontsize=7.5)
                add_text((225.5 - w_ta, y_ta), f"{t_val}/{a_val}", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((228.2, y_ta), "°", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                w_nm = pymupdf.get_text_length(str(mods_cnt), fontname="helv", fontsize=7.5)
                add_text((225.5 - w_nm, y_nm), str(mods_cnt), fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((228.2, y_nm), "units", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                w_stc = pymupdf.get_text_length(f"{stc_v:.0f}", fontname="helv", fontsize=7.5)
                add_text((225.5 - w_stc, y_stc), f"{stc_v:.0f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((228.2, y_stc), "Wp", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                mod_str = f"1 unit x {mods_cnt}"
                w_mod = pymupdf.get_text_length(mod_str, fontname="helv", fontsize=7.5)
                add_text((225.2 - w_mod, y_mod), mod_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((228.2, y_mod), "In series", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                w_pm = pymupdf.get_text_length(f"{pmpp_v:.0f}", fontname="helv", fontsize=7.5)
                add_text((225.5 - w_pm, y_pmpp), f"{pmpp_v:.0f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((228.2, y_pmpp), "Wp", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                um_str = f"{umpp_v:.1f}" if isinstance(umpp_v, float) and not umpp_v.is_integer() else f"{int(umpp_v)}"
                w_um = pymupdf.get_text_length(um_str, fontname="helv", fontsize=7.5)
                add_text((225.5 - w_um, y_umpp), um_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((228.2, y_umpp), "V", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                w_im = pymupdf.get_text_length(f"{impp_v:.0f}", fontname="helv", fontsize=7.5)
                add_text((225.5 - w_im, y_impp), f"{impp_v:.0f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((228.2, y_impp), "A", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                # Right Inverter side
                add_redact(pymupdf.Rect(380, r_arr.y0 + 25, 545, r_arr.y0 + 110))
                add_text((426, y_ni), f"1 * MPPT {mppt_pct}  {inv_frac}", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((475, y_tp), str(tot_p), fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((465, y_op), str(op_v), fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((479, y_pr), str(pnom_r), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            render_array_sec("Array #1 - PV Array", arrays_data[0] if len(arrays_data) > 0 else None, 0, "20/103°")
            render_array_sec("Array #2 - Sub-array #2", arrays_data[1] if len(arrays_data) > 1 else None, 1, "20/-167°")
            render_array_sec("Array #3 - Sub-array #3", arrays_data[2] if len(arrays_data) > 2 else None, 2, "20/13°")
            render_array_sec("Array #4 - Sub-array #4", arrays_data[3] if len(arrays_data) > 3 else None, 3, "20/-77°")

            # Total PV power & Total inverter power
            tpv_m = page.search_for("Total PV power")
            if tpv_m:
                r_tpv = tpv_m[0]
                add_redact(pymupdf.Rect(180, r_tpv.y0 + 8, 265, r_tpv.y0 + 22))
                add_text((200, r_tpv.y0 + 20), f"{variant_str} kWp", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_redact(pymupdf.Rect(180, r_tpv.y0 + 23, 265, r_tpv.y0 + 34))
                add_text((195, r_tpv.y0 + 32), f"{total_mods} modules", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                
                mod_area_val = total_mods * 2.7
                cell_area_val = total_mods * 2.52
                add_redact(pymupdf.Rect(180, r_tpv.y0 + 35, 265, r_tpv.y0 + 46))
                add_text((205, r_tpv.y0 + 44), f"{mod_area_val:.1f} m²", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_redact(pymupdf.Rect(180, r_tpv.y0 + 47, 265, r_tpv.y0 + 58))
                add_text((205, r_tpv.y0 + 56), f"{cell_area_val:.1f} m²", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            tinv_m = page.search_for("Total inverter power")
            if tinv_m:
                r_tinv = tinv_m[0]
                add_redact(pymupdf.Rect(460, r_tinv.y0 + 8, 545, r_tinv.y0 + 55))
                add_text((475, r_tinv.y0 + 20), f"{kwac_str} kWac", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((485, r_tinv.y0 + 32), f"{inv_units} unit", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((479, r_tinv.y0 + 44), f"{pnom_rat:.2f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

        # 8. Dynamic Grid Storage (p_idx >= 2)
        if p_idx >= 2 and "Grid storage" in p_txt and has_bat:
            grid_st_m = page.search_for("Grid storage")
            if grid_st_m:
                r_gst = grid_st_m[0]
                add_redact(pymupdf.Rect(175, r_gst.y0 + 24, 240, r_gst.y0 + 36))
                add_text((192, r_gst.y0 + 34), "Hoymiles", fontsize=7.5, fontname="helv", color=(0, 0, 0))
                
                add_redact(pymupdf.Rect(175, r_gst.y0 + 36, 260, r_gst.y0 + 48))
                add_text((180, r_gst.y0 + 46), bat_model, fontsize=7.5, fontname="helv", color=(0, 0, 0))

                add_redact(pymupdf.Rect(175, r_gst.y0 + 60, 245, r_gst.y0 + 72))
                add_text((200, r_gst.y0 + 70), f"{bat_units} in parallel", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                add_redact(pymupdf.Rect(195, r_gst.y0 + 72, 235, r_gst.y0 + 84))
                add_text((205, r_gst.y0 + 82), f"{bat_volt} V", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                add_redact(pymupdf.Rect(195, r_gst.y0 + 84, 240, r_gst.y0 + 96))
                add_text((205, r_gst.y0 + 94), f"{bat_cap} Ah", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                add_redact(pymupdf.Rect(460, r_gst.y0 + 36, 535, r_gst.y0 + 48))
                add_text((475, r_gst.y0 + 46), "8.01 kWdc", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                add_redact(pymupdf.Rect(460, r_gst.y0 + 84, 535, r_gst.y0 + 96))
                add_text((475, r_gst.y0 + 94), "7.61 kWac", fontsize=7.5, fontname="helv", color=(0, 0, 0))

        # 9. DC & AC Wiring Losses (p_idx >= 2)
        if p_idx >= 2 and "DC wiring losses" in p_txt:
            dc_m = page.search_for("DC wiring losses")
            if dc_m:
                r_dc = dc_m[0]
                lf_matches = [r for r in page.search_for("Loss Fraction") if r_dc.y0 < r.y0 < r_dc.y0 + 50 and r.x0 < 100]
                if lf_matches:
                    r_lf = lf_matches[0]
                    gl_val = dc_l.get("global_loss", 2.0)
                    w_gl = pymupdf.get_text_length(f"{gl_val:.2f}", fontname="helv", fontsize=7.5)
                    add_redact(pymupdf.Rect(130, r_lf.y0 - 2, 160.5, r_lf.y1 + 2))
                    add_text((159.0 - w_gl, r_lf.y1 - 1), f"{gl_val:.2f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                w_res1 = pymupdf.get_text_length(str(dc_l["arr1_res"]), fontname="helv", fontsize=7.5)
                add_redact(pymupdf.Rect(200, r_dc.y0 + 58, 226.5, r_dc.y0 + 70))
                add_text((225.5 - w_res1, r_dc.y0 + 67), str(dc_l["arr1_res"]), fontsize=7.5, fontname="helv", color=(0, 0, 0))

                w_los1 = pymupdf.get_text_length(f"{dc_l['arr1_loss']:.2f}", fontname="helv", fontsize=7.5)
                add_redact(pymupdf.Rect(200, r_dc.y0 + 70, 226.5, r_dc.y0 + 82))
                add_text((225.5 - w_los1, r_dc.y0 + 79), f"{dc_l['arr1_loss']:.2f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                w_res2 = pymupdf.get_text_length(str(dc_l["arr2_res"]), fontname="helv", fontsize=7.5)
                add_redact(pymupdf.Rect(460, r_dc.y0 + 58, 497.0, r_dc.y0 + 70))
                add_text((495.5 - w_res2, r_dc.y0 + 67), str(dc_l["arr2_res"]), fontsize=7.5, fontname="helv", color=(0, 0, 0))

                w_los2 = pymupdf.get_text_length(f"{dc_l['arr2_loss']:.2f}", fontname="helv", fontsize=7.5)
                add_redact(pymupdf.Rect(460, r_dc.y0 + 70, 497.0, r_dc.y0 + 82))
                add_text((495.5 - w_los2, r_dc.y0 + 79), f"{dc_l['arr2_loss']:.2f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                if "Array #3 - Sub-array #3" in p_txt:
                    w_res3 = pymupdf.get_text_length(str(dc_l["arr3_res"]), fontname="helv", fontsize=7.5)
                    add_redact(pymupdf.Rect(200, r_dc.y0 + 94, 226.5, r_dc.y0 + 106))
                    add_text((225.5 - w_res3, r_dc.y0 + 103.3), str(dc_l["arr3_res"]), fontsize=7.5, fontname="helv", color=(0, 0, 0))

                    w_los3 = pymupdf.get_text_length(f"{dc_l['arr3_loss']:.2f}", fontname="helv", fontsize=7.5)
                    add_redact(pymupdf.Rect(200, r_dc.y0 + 106, 226.5, r_dc.y0 + 118))
                    add_text((225.5 - w_los3, r_dc.y0 + 115.3), f"{dc_l['arr3_loss']:.2f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                if "Array #4 - Sub-array #4" in p_txt and dc_l.get("arr4_res", 0) > 0:
                    w_res4 = pymupdf.get_text_length(str(dc_l["arr4_res"]), fontname="helv", fontsize=7.5)
                    add_redact(pymupdf.Rect(460, r_dc.y0 + 94, 497.0, r_dc.y0 + 106))
                    add_text((495.5 - w_res4, r_dc.y0 + 103.3), str(dc_l["arr4_res"]), fontsize=7.5, fontname="helv", color=(0, 0, 0))

                    w_los4 = pymupdf.get_text_length(f"{dc_l['arr4_loss']:.2f}", fontname="helv", fontsize=7.5)
                    add_redact(pymupdf.Rect(460, r_dc.y0 + 106, 497.0, r_dc.y0 + 118))
                    add_text((495.5 - w_los4, r_dc.y0 + 115.3), f"{dc_l['arr4_loss']:.2f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

        if p_idx >= 2 and "AC wiring losses" in p_txt:
            ac_m = page.search_for("AC wiring losses")
            if ac_m:
                r_ac = ac_m[0]
                add_redact(pymupdf.Rect(200, r_ac.y0 + 20, 280, r_ac.y0 + 36))
                add_text((200, r_ac.y0 + 33), ac_l["inverter_voltage"], fontsize=7.5, fontname="helv", color=(0, 0, 0))

                add_redact(pymupdf.Rect(200, r_ac.y0 + 36, 280, r_ac.y0 + 48))
                add_text((195, r_ac.y0 + 45), "1.00 % at STC", fontsize=7.5, fontname="helv", color=(0, 0, 0))

                add_redact(pymupdf.Rect(170, r_ac.y0 + 58, 275, r_ac.y0 + 72))
                add_text((179, r_ac.y0 + 69), ac_l["wire_section"], fontsize=7.5, fontname="helv", color=(0, 0, 0))

                wlen_num = str(ac_l.get("wires_length", "10 m")).replace("m", "").strip()
                w_wlen = pymupdf.get_text_length(wlen_num, fontname="helv", fontsize=7.5)
                add_redact(pymupdf.Rect(190, r_ac.y0 + 72, 275, r_ac.y0 + 86))
                add_text((225.5 - w_wlen, r_ac.y0 + 81), wlen_num, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((228.2, r_ac.y0 + 81), "m", fontsize=7.5, fontname="helv", color=(0, 0, 0))

        # 10. Main results (p_idx >= 3)
        if p_idx >= 3 and "Balances and main results" in p_txt:
            pe_m2 = page.search_for("Produced Energy")
            if pe_m2:
                per2 = pe_m2[0]
                add_redact(pymupdf.Rect(160, per2.y0 - 2, 227, per2.y1 + 2))
                w_pe2 = pymupdf.get_text_length(str(prod_kwh), fontname="helv", fontsize=7.5)
                add_text((226 - w_pe2, per2.y1 - 1), str(prod_kwh), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            ue_m2 = page.search_for("Used Energy")
            if ue_m2 and has_bat:
                uer2 = ue_m2[0]
                add_redact(pymupdf.Rect(160, uer2.y0 - 2, 227, uer2.y1 + 2))
                w_ue2 = pymupdf.get_text_length(str(used_kwh), fontname="helv", fontsize=7.5)
                add_text((226 - w_ue2, uer2.y1 - 1), str(used_kwh), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            sp_m2 = page.search_for("Specific prod.") or page.search_for("Specific production")
            if sp_m2:
                spr2 = sp_m2[0]
                add_redact(pymupdf.Rect(420, spr2.y0 - 2, 496, spr2.y1 + 2))
                w_sp2 = pymupdf.get_text_length(spec_str, fontname="helv", fontsize=7.5)
                add_text((495 - w_sp2, spr2.y1 - 1), spec_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))

            pr_m2 = page.search_for("Perf. Ratio")
            if pr_m2:
                prr2 = pr_m2[0]
                add_redact(pymupdf.Rect(420, prr2.y0 - 2, 496, prr2.y1 + 2))
                w_pr2 = pymupdf.get_text_length(f"{pr_pct:.2f}", fontname="helv", fontsize=7.5)
                add_text((495 - w_pr2, prr2.y1 - 1), f"{pr_pct:.2f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            sf_m2 = page.search_for("Solar Fraction")
            if sf_m2 and has_bat:
                sfr2 = sf_m2[0]
                add_redact(pymupdf.Rect(420, sfr2.y0 - 2, 496, sfr2.y1 + 2))
                w_sf2 = pymupdf.get_text_length(f"{sf_pct:.2f}", fontname="helv", fontsize=7.5)
                add_text((495 - w_sf2, sfr2.y1 - 1), f"{sf_pct:.2f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            # PR Chart Legend: PR: Performance Ratio (Yf / Yr) : 0.8145
            pr_chart_str = f"{pr_pct / 100.0:.4f}"
            for w in page.get_text("words"):
                if 220 < w[1] < 340 and 420 < w[0] < 500 and ("0." in w[4] or w[4] == "0.737"):
                    add_redact(pymupdf.Rect(w[0] - 1, w[1] - 1, w[2] + 18, w[3] + 1))
                    add_text((w[0], w[3] - 1), pr_chart_str, fontsize=6.8, fontname="helv", color=(0, 0, 0))

            # Table row balances
            bal_matrix = excel_data.get("balances_table") or excel_data.get("main_results_table", [])
            if bal_matrix and len(bal_matrix) >= 13:
                if has_bat:
                    col_x_ends = [142.0, 185.0, 227.0, 273.0, 317.0, 363.0, 408.0, 453.0, 498.0, 545.0]
                    for r_idx in range(13):
                        y_curr = 527.0 + (r_idx * 12.0) if r_idx < 12 else 674.5
                        add_redact(pymupdf.Rect(95, y_curr - 8, 550, y_curr + 4))
                        r_data = bal_matrix[r_idx]
                        for c_idx in range(min(10, len(r_data))):
                            val_str = str(r_data[c_idx])
                            w_val = pymupdf.get_text_length(val_str, fontname="helv", fontsize=7.0)
                            add_text((col_x_ends[c_idx] - w_val, y_curr), val_str, fontsize=7.0, fontname="helv", color=(0, 0, 0))
                else:
                    col_x_ends_nb = [160.0, 215.0, 270.0, 325.0, 380.0, 435.0, 490.0, 545.0]
                    for r_idx in range(13):
                        y_curr = 466.0 + (r_idx * 12.0) if r_idx < 12 else 614.5
                        add_redact(pymupdf.Rect(110, y_curr - 8, 550, y_curr + 4))
                        r_data = bal_matrix[r_idx]
                        nobat_cols = [
                            r_data[0] if len(r_data) > 0 else "0",
                            r_data[1] if len(r_data) > 1 else "0",
                            r_data[2] if len(r_data) > 2 else "0",
                            r_data[3] if len(r_data) > 3 else "0",
                            r_data[4] if len(r_data) > 4 else "0",
                            r_data[5] if len(r_data) > 5 else "0",
                            r_data[8] if len(r_data) > 8 else (r_data[5] if len(r_data) > 5 else "0"),
                            f"{pr_pct/100:.3f}" if r_idx == 12 else "0.780"
                        ]
                        for c_idx in range(min(8, len(nobat_cols))):
                            val_str = str(nobat_cols[c_idx])
                            w_val = pymupdf.get_text_length(val_str, fontname="helv", fontsize=7.0)
                            add_text((col_x_ends_nb[c_idx] - w_val, y_curr), val_str, fontsize=7.0, fontname="helv", color=(0, 0, 0))

        # 11. Loss diagram (p_idx >= 3)
        if p_idx >= 3 and "Loss diagram" in p_txt and ("PV nominal" in p_txt or "Array virtual energy" in p_txt or "PV conversion" in p_txt):
            # Box 1: Global horizontal (y: 135..155)
            add_redact(pymupdf.Rect(120, 138, 168.0, 154))
            gh_str = f"{gh_v:.1f}"
            add_text((167.0 - pymupdf.get_text_length(gh_str, fontname="helv", fontsize=7.5), 150), gh_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))

            # Box 2: Effective irradiation (y: 215..232)
            add_redact(pymupdf.Rect(120, 218, 162.0, 230))
            gi_str = f"{gi_v:.0f}" if gi_v.is_integer() else f"{gi_v:.1f}"
            add_text((160.0 - pymupdf.get_text_length(gi_str, fontname="helv", fontsize=7.5), 226.5), gi_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))

            # Collector area below effective irradiation: "* 38 m² coll."
            add_redact(pymupdf.Rect(120, 231, 195.0, 245))
            coll_str = f"* {round(total_mods * 2.7)} m² coll."
            add_text((140, 240.0), coll_str, fontsize=7.0, fontname="helv", color=(0, 0, 0))

            # Box 3: PV nominal energy (y: 268..285)
            add_redact(pymupdf.Rect(120, 270, 170.0, 282))
            pvn_str = f"{pvn_v:.1f}"
            add_text((169.0 - pymupdf.get_text_length(pvn_str, fontname="helv", fontsize=7.5), 278.5), pvn_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))

            # Box 4: Virtual energy (y: 395..425)
            add_redact(pymupdf.Rect(105, 400, 195.0, 424))
            av_str = f"{av_v:.0f}"
            add_text((130.0, 409.0), f"{av_str} kWh", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            # Box 5: Available energy (y: 515..545)
            add_redact(pymupdf.Rect(105, 520, 195.0, 545))
            fe_str = f"{fe_v:.0f}"
            add_text((130.0, 528.5), f"{fe_str} kWh", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            if has_bat:
                # Bottom 3 boxes (clean integer formatting, right-aligned within box bounds)
                add_redact(pymupdf.Rect(55, 698, 92, 712))
                add_redact(pymupdf.Rect(108, 698, 145, 712))
                add_redact(pymupdf.Rect(165, 698, 204, 712))

                ef_str = f"{int(round(ef_v))}"
                bs_str = f"{int(round(bs_v))}"
                du_str = f"{int(round(du_v))}"

                w_ef = pymupdf.get_text_length(ef_str, fontname="helv", fontsize=7.5)
                w_bs = pymupdf.get_text_length(bs_str, fontname="helv", fontsize=7.5)
                w_du = pymupdf.get_text_length(du_str, fontname="helv", fontsize=7.5)

                add_text((83.5 - w_ef, 707.5), ef_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((131.0 - w_bs, 707.5), bs_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((193.0 - w_du, 707.5), du_str, fontsize=7.5, fontname="helv", color=(0, 0, 0))

                # Percentages above bottom boxes
                add_redact(pymupdf.Rect(50, 606, 95, 618))
                add_redact(pymupdf.Rect(90, 635, 125, 648))
                add_redact(pymupdf.Rect(150, 635, 185, 648))

                pct_g = (ef_v / used_kwh * 100.0) if used_kwh > 0 else 0.0
                tot_disp = bs_v + du_v
                pct_st = (bs_v / tot_disp * 100.0) if tot_disp > 0 else 50.0
                pct_du = (du_v / tot_disp * 100.0) if tot_disp > 0 else 50.0

                str_pg = f"{pct_g:.1f}%"
                str_pst = f"{pct_st:.1f}%"
                str_pdu = f"{pct_du:.1f}%"

                add_text((82.0 - pymupdf.get_text_length(str_pg, fontname="helv", fontsize=7.2), 616.0), str_pg, fontsize=7.2, fontname="helv", color=(0, 0, 0))
                add_text((120.0 - pymupdf.get_text_length(str_pst, fontname="helv", fontsize=7.2), 645.0), str_pst, fontsize=7.2, fontname="helv", color=(0, 0, 0))
                add_text((180.0 - pymupdf.get_text_length(str_pdu, fontname="helv", fontsize=7.2), 645.0), str_pdu, fontsize=7.2, fontname="helv", color=(0, 0, 0))
            else:
                add_redact(pymupdf.Rect(105, 595, 175.0, 615))
                str_egr = f"{int(round(fe_v))}"
                w_egr = pymupdf.get_text_length(str_egr, fontname="helv", fontsize=7.5)
                add_text((130.0, 606.0), str_egr, fontsize=7.5, fontname="helv", color=(0, 0, 0))
                add_text((130.0 + w_egr + 3, 606.0), "kWh", fontsize=7.5, fontname="helv", color=(0, 0, 0))

        # 12. P50 - P90 evaluation (p_idx >= 3)
        if p_idx >= 3 and "P50 - P90 evaluation" in p_txt and "Annual production probability" in p_txt:
            # Word-based redaction for top table numbers
            for w in page.get_text("words"):
                if 180 < w[1] < 290 and (w[0] > 480 or (220 < w[0] < 250)):
                    if any(c.isdigit() for c in w[4]) and w[4] not in ["0.0", "1.0", "5.9"]:
                        r_max_x = min(w[2] + 0.5, 513.5) if w[0] > 480 else (w[2] + 1)
                        add_redact(pymupdf.Rect(w[0] - 1, w[1] - 1, r_max_x, w[3] + 1))

            w_var = pymupdf.get_text_length(f"{var_v:.1f}", fontname="helv", fontsize=7.5)
            add_text((242.4 - w_var, 196.5), f"{var_v:.1f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            w_cvar = pymupdf.get_text_length(f"{cvar_v:.1f}", fontname="helv", fontsize=7.5)
            add_text((511.7 - w_cvar, 208.5), f"{cvar_v:.1f}", fontsize=7.5, fontname="helv", color=(0, 0, 0))

            w_pv = pymupdf.get_text_length(str(p_var_n), fontname="helv", fontsize=7.5)
            add_text((513.0 - w_pv, 250.5), str(p_var_n), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            w_p50 = pymupdf.get_text_length(str(p50_n), fontname="helv", fontsize=7.5)
            add_text((513.0 - w_p50, 262.5), str(p50_n), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            w_p90 = pymupdf.get_text_length(str(p90_n), fontname="helv", fontsize=7.5)
            add_text((513.0 - w_p90, 274.5), str(p90_n), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            w_p75 = pymupdf.get_text_length(str(p75_n), fontname="helv", fontsize=7.5)
            add_text((513.0 - w_p75, 286.5), str(p75_n), fontsize=7.5, fontname="helv", color=(0, 0, 0))

            # Dynamic word-based redaction and positioning for graph numbers across all 6 templates
            p50_pos, egrid_pos, p75_pos, p90_pos = None, None, None, None
            for w in page.get_text("words"):
                if 340 < w[1] < 520 and 230 < w[0] < 450 and any(c.isdigit() for c in w[4]):
                    add_redact(pymupdf.Rect(w[0] - 2, w[1] - 2, w[2] + 2, w[3] + 2))
                    if 345 < w[1] < 370 and w[0] < 345:
                        p50_pos = (w[0], w[3] - 1)
                    elif 355 < w[1] < 380 and w[0] >= 345:
                        egrid_pos = (w[0], w[3] - 1)
                    elif 395 < w[1] < 430:
                        p75_pos = (w[0], w[3] - 1)
                    elif 465 < w[1] < 505:
                        p90_pos = (w[0], w[3] - 1)

            if not p50_pos: p50_pos = (304.0, 363.5)
            if not egrid_pos: egrid_pos = (372.0, 372.5)
            if not p75_pos: p75_pos = (303.0, 415.0)
            if not p90_pos: p90_pos = (267.5, 490.0)

            add_text(p50_pos, str(p50_n), fontsize=7.2, fontname="helv", color=(0, 0.45, 0))
            add_text(egrid_pos, str(egrid_n), fontsize=7.2, fontname="helv", color=(0.85, 0.45, 0))
            add_text(p75_pos, str(p75_n), fontsize=7.2, fontname="helv", color=(0, 0.45, 0))
            add_text(p90_pos, str(p90_n), fontsize=7.2, fontname="helv", color=(0, 0.45, 0))

        # EXECUTE CLEAN 2-PHASE UPDATE FOR THIS PAGE:
        # Phase A: Add all redaction annotations with fill=False (removes text characters without drawing white boxes!)
        for r in redacts:
            page.add_redact_annot(r, fill=False)
            
        # Delete highlight & FreeText annotations (removes yellow highlight boxes and template annotations!)
        for a in list(page.annots()):
            if a.type[0] in [2, 8]:
                page.delete_annot(a)

        # Phase B: Apply redactions (purges text glyphs without painting white boxes!)
        page.apply_redactions(images=0, graphics=0)
        
        # Phase C: Insert all new text onto pristine page canvas
        for pos, text, fsize, fname, clr in inserts:
            page.insert_text(pos, text, fontsize=fsize, fontname=fname, color=clr)

    pdf_bytes = doc.tobytes()
    doc.close()
    return pdf_bytes
